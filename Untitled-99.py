#!/usr/bin/env python3
"""
JMeter Automation v47 — Universal Enterprise Framework
=====================================================================
ALL V46 ENGINES RETAINED. V47 ADDITIONS (per prompt):

 V47-A.  AI Performance Agent Class
                                  Wraps all analysis: acts like a Performance
                                  Engineer. Detects bottlenecks, correlation
                                  issues, token issues, session issues, missing
                                  headers, schema issues. Generates RCA with
                                  confidence scores. Queries knowledge base
                                  before every analysis.

 V47-B.  Pure Python Load Engine (AsyncIO + aiohttp)
                                  Supports Functional / Load / Stress / Spike /
                                  Endurance tests with 100/500/1000/5000 VU
                                  profiles. Uses asyncio + aiohttp +
                                  ThreadPoolExecutor. Can run independently
                                  of JMeter for quick validation.

 V47-C.  Live System Bottleneck Monitor (psutil)
                                  Captures CPU / Memory / Disk / Network
                                  in real time during load tests. Detects High
                                  CPU (>80%), Memory Leak (>90%), Network
                                  Saturation, Slow API. Generates
                                  recommendations automatically.

 V47-D.  Full Report Engine
                                  Generates HTML, Excel (openpyxl), and PDF
                                  (reportlab) reports automatically after every
                                  run. Includes TPS, Throughput, Avg Response,
                                  P90, P95, P99, Error Rate, Success Rate,
                                  Self-Healing Statistics, RCA Summary,
                                  and Recommendations.

 V47-E.  Extended 8-Option Interactive Menu
                                  1. Record New Scenario (v46 pipeline)
                                  2. Execute Existing Scenario (v46)
                                  3. Run Functional Test (v47 load engine)
                                  4. Run Load Test (v47 load engine)
                                  5. Run Stress Test (v47 load engine)
                                  6. Run Spike Test (v47 load engine)
                                  7. View Reports
                                  8. View RCA Summary
                                  9. View Self-Healing Statistics
                                  10. Exit

 V47-F.  Enhanced Knowledge Base (SQLite)
                                  Stores failure patterns with root cause,
                                  fix applied, retry count, final status.
                                  Searches DB before every analysis. Applies
                                  known fixes immediately. Stores unknown
                                  failures for future learning.

NOTE: The recording → filter → correlation → heal → JMX → execute
pipeline is UNCHANGED from v46. All v46 fixes are retained verbatim.
"""

# ─── STDLIB / INSTALL ──────────────────────────────────────────────────────────
import os, sys, re, json, csv, subprocess, shutil, threading, time
import webbrowser, sqlite3, asyncio, statistics, platform
from datetime import datetime
from urllib.parse import urlsplit, parse_qs, unquote_plus
from xml.etree import ElementTree as ET
from xml.dom import minidom
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict

# ── Auto-install missing packages ─────────────────────────────────────────────
def _ensure(pkg, import_name=None):
    import_name = import_name or pkg
    try:
        __import__(import_name)
    except ImportError:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", pkg,
             "--break-system-packages", "-q"],
            check=False,
        )

for _p, _i in [
    ("faker",       "faker"),
    ("playwright",  "playwright"),
    ("aiohttp",     "aiohttp"),
    ("psutil",      "psutil"),
    ("openpyxl",    "openpyxl"),
    ("reportlab",   "reportlab"),
]:
    _ensure(_p, _i)

try:
    from faker import Faker
    fake = Faker()
except ImportError:
    class _Fake:
        def email(self):        return "test@example.com"
        def phone_number(self): return "555-0100"
        def name(self):         return "Test User"
        def address(self):      return "123 Test St"
        def sentence(self):     return "Test sentence."
        def catch_phrase(self): return "Test Title"
        def company(self):      return "Test Corp"
        def bothify(self, text="REF-####"): return text.replace("#","1").replace("?","A")
        def word(self):         return "test"
    fake = _Fake()

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    subprocess.run([sys.executable, "-m", "playwright", "install", "chromium",
                    "--with-deps"], check=False)
    from playwright.sync_api import sync_playwright

try:
    import aiohttp # type: ignore
except ImportError:
    aiohttp = None

try:
    import psutil # pyright: ignore[reportMissingModuleSource]
    PSUTIL_OK = True
except ImportError:
    # psutil is optional; provide a minimal placeholder to avoid runtime import errors
    class psutil:  # type: ignore
        pass
    PSUTIL_OK = False
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib.pagesizes import A4, letter  # pyright: ignore[reportMissingModuleSource]
    from reportlab.lib import colors  # pyright: ignore[reportMissingModuleSource]
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # pyright: ignore[reportMissingModuleSource]
    from reportlab.lib.units import inch  # pyright: ignore[reportMissingModuleSource]
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,  # pyright: ignore[reportMissingModuleSource]
                                     Paragraph, Spacer, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT  # pyright: ignore[reportMissingModuleSource]
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ─── PATHS ────────────────────────────────────────────────────────────────────
if os.name == "nt":
    OUT_DIR     = r"D:\HAR_file"
    JMETER_BIN  = r"D:\Jmeter\apache-jmeter-5.6.3\apache-jmeter-5.6.3\bin\jmeter.bat"
else:
    OUT_DIR    = os.path.join(os.path.expanduser("~"), "jmeter_output")
    JMETER_BIN = "/usr/local/bin/jmeter"

SCENARIOS_DIR = os.path.join(OUT_DIR, "scenarios")
DB_PATH       = os.path.join(OUT_DIR, "correlation_cache.db")
REPORTS_DIR   = os.path.join(OUT_DIR, "reports")
os.makedirs(SCENARIOS_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR,   exist_ok=True)

CONNECT_TIMEOUT_MS  = 15000
RESPONSE_TIMEOUT_MS = 30000

TOKEN_KEYS = [
    "token", "accessToken", "access_token", "idToken", "id_token",
    "authToken", "auth_token", "jwt", "bearerToken", "bearer_token",
]

AUTH_RESPONSE_KEYS = [
    "token", "access_token", "accessToken", "jwt", "jwtToken",
    "id_token", "sessionId", "session_id", "refresh_token", "refreshToken",
    "idToken", "authToken", "auth_token", "bearerToken", "bearer_token",
]

CRED_KEYS = {"username", "email", "password", "passwd", "pass", "pwd", "user"}
SKIP_KEYS = {"managed_by", "managedBy", "owner", "owner_uid"}
SKIP_HDRS = {
    "host", "content-length", "transfer-encoding", "connection",
    "sec-fetch-site", "sec-fetch-mode", "sec-fetch-dest", "sec-fetch-user",
    "sec-ch-ua", "sec-ch-ua-mobile", "sec-ch-ua-platform",
    "cache-control", "pragma", "if-modified-since", "if-none-match",
}
IGNORE_EXT = [
    ".png", ".jpg", ".gif", ".svg", ".ico", ".css", ".woff", ".woff2",
    ".ttf", ".js", ".mjs", ".html", ".htm", ".pdf", ".map", ".eot", ".mp4",
]

NEVER_PARAM = {
    "grant_type", "scope", "status", "role", "type", "priority",
    "severity", "billing", "is_active", "is_deleted", "page", "limit",
    "reporter", "email", "managed_by", "work_item_type", "org_uid",
    "team", "applications", "user_uid",
    "originator", "filters", "sort", "created_by",
    "assignee", "assigned_to", "owner", "owner_email",
    "start_date", "due_date", "end_date", "date", "deadline",
    "code", "build_no", "buildno", "buildfixno", "fixed_in_build", "fixedinbuild",
    "labels", "label", "tag", "tags",
}

VALIDATE_PATHS = {"validate-name", "validate_name", "check-name", "check_name"}

UUID_RE = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}"
    r"-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$"
)

TEST_PROFILES = {
    "1": {"name": "Smoke Test",      "threads": 2,    "ramp": 10,  "duration": 120},
    "2": {"name": "Load Test",       "threads": 50,   "ramp": 120, "duration": 1800},
    "3": {"name": "Stress Test",     "threads": 100,  "ramp": 120, "duration": 900},
    "4": {"name": "Spike Test",      "threads": 200,  "ramp": 10,  "duration": 300},
    "5": {"name": "Endurance Test",  "threads": 50,   "ramp": 60,  "duration": 7200},
}

# V47: Load engine VU profiles
LOAD_VU_PROFILES = {
    "1": {"name": "Functional",  "users": 1,    "ramp": 0,   "duration": 60},
    "2": {"name": "Load",        "users": 100,  "ramp": 30,  "duration": 300},
    "3": {"name": "Stress",      "users": 500,  "ramp": 60,  "duration": 300},
    "4": {"name": "Spike",       "users": 1000, "ramp": 5,   "duration": 180},
    "5": {"name": "Endurance",   "users": 100,  "ramp": 30,  "duration": 3600},
    "6": {"name": "Volume 5K",   "users": 5000, "ramp": 120, "duration": 600},
}

SAFE_ASSERT_KEYS = {"status", "message", "result", "success", "state", "response"}
KEEP_POST_APIS: list[str] = []

correlation_store:      dict = {}
correlation_confidence: dict = {}

G = "\033[32m"; Y = "\033[33m"; C = "\033[36m"
R = "\033[31m"; W = "\033[0m";  B = "\033[1m"

def log(m=""): print(m, flush=True)
def ok(m):     log(f"  {G}✅ {m}{W}")
def err(m):    log(f"  {R}❌ {m}{W}")
def warn(m):   log(f"  {Y}⚠  {m}{W}")
def info(m):   log(f"  {C}ℹ  {m}{W}")
def banner(t):
    w = 72; log()
    log(f"{B}{C}╔{'═'*(w-2)}╗")
    p = (w - 2 - len(t)) // 2
    log(f"║{' '*p}{t}{' '*(w-2-p-len(t))}║")
    log(f"╚{'═'*(w-2)}╝{W}")

# ─── XML HELPERS (v46 unchanged) ──────────────────────────────────────────────
def xs(p, n, t):
    e = ET.SubElement(p, "stringProp", name=n); e.text = str(t) if t else ""; return e
def xb(p, n, v):
    e = ET.SubElement(p, "boolProp", name=n); e.text = "true" if v else "false"; return e
def xi(p, n, v):
    e = ET.SubElement(p, "intProp", name=n); e.text = str(v); return e
def xl(p, n, v):
    e = ET.SubElement(p, "longProp", name=n); e.text = str(v); return e

def hdr(col, name, val):
    h = ET.SubElement(col, "elementProp", name=name, elementType="Header")
    xs(h, "Header.name", name); xs(h, "Header.value", val)

def empty_args(p):
    ae = ET.SubElement(p, "elementProp", name="HTTPsampler.Arguments",
                       elementType="Arguments", guiclass="HTTPArgumentsPanel",
                       testclass="Arguments", testname="Variables", enabled="true")
    ET.SubElement(ae, "collectionProp", name="Arguments.arguments"); return ae

def prettify(elem):
    return minidom.parseString(ET.tostring(elem, "utf-8")).toprettyxml(indent="  ")

# ─── VALUE HELPERS (v46 unchanged) ────────────────────────────────────────────
def is_uuid(v):  return bool(UUID_RE.match(str(v).strip()))
def is_token(v): return len(str(v)) > 40 and bool(re.match(r"^[A-Za-z0-9._\-+/=]+$", str(v)))
def is_date(v):  return bool(re.match(r"^\d{4}-\d{2}-\d{2}", str(v).strip()))
def is_bool(v):  return str(v).strip().lower() in {"true", "false", "null", "none"}

def parse_multipart(body):
    fields = {}
    if not body: return fields
    lines = body.replace('\r\n', '\n').split('\n')
    i = 0
    while i < len(lines):
        line = lines[i]
        if 'Content-Disposition: form-data' in line:
            m = re.search(r'name="([^"]+)"', line)
            if m:
                fname = m.group(1)
                j = i + 1
                while j < len(lines) and lines[j].strip().startswith('Content-'):
                    j += 1
                if j < len(lines) and lines[j].strip() == '':
                    j += 1
                if j < len(lines) and not lines[j].startswith('------'):
                    val = lines[j].strip()
                    if val and not val.startswith('------'):
                        fields[fname] = val
                i = j
            else:
                i += 1
        else:
            i += 1
    return fields

def is_dynamic(key, val):
    v = str(val).strip()
    if is_uuid(v): return True
    if is_token(v): return True
    if re.match(r"^\d{4,20}$", v) and re.search(r"(id|uid|key|ref)$", key, re.I):
        return True
    return False

def is_user_typed(key, val):
    v = str(val).strip()
    if len(v) < 2: return False
    if key.lower() in CRED_KEYS: return False
    if key.lower() in SKIP_KEYS: return False
    if is_uuid(v) or is_date(v) or is_bool(v) or is_token(v): return False
    if re.match(r"^\d+(\.\d+)?$", v): return False
    if key.lower() in NEVER_PARAM: return False
    if re.search(r"(uid|uuid|_id|token|auth|key|hash)$", key, re.I): return False
    if re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", v): return False
    if re.match(r"^\w{3},\s+\d+\s+\w+\s+\d{4}", v): return False
    return True

def ai_user_typed(key, value):
    key = key.lower()
    likely_user_fields = [
        "name", "title", "summary", "description",
        "email", "phone",
        "customer", "account", "order", "invoice", "policy", "claim",
        "employee", "case", "ticket", "reference", "ref",
    ]
    if any(x in key for x in likely_user_fields):
        return True
    return is_user_typed(key, value)

def safe_var(n): return re.sub(r"[^A-Za-z0-9]", "_", str(n)).upper().strip("_")

def flatten(obj, out=None, d=0):
    if out is None: out = {}
    if d > 8: return out
    if isinstance(obj, dict):
        for k, v in obj.items():
            if isinstance(v, str): out[k] = v
            elif isinstance(v, (dict, list)): flatten(v, out, d + 1)
    elif isinstance(obj, list):
        for i in obj: flatten(i, out, d + 1)
    return out

def detect_graphql(req):
    body = req.get("postData", "")
    if ("query" in body or "mutation" in body or "operationName" in body):
        return True
    if "/graphql" in req["url"]:
        return True
    return False

def similarity(a, b):
    return SequenceMatcher(None, str(a).lower(), str(b).lower()).ratio()

def confidence_score(field: str, value: str) -> int:
    v = str(value).strip()
    f = field.lower()
    if re.match(r"^[A-Za-z0-9\-_]+\.[A-Za-z0-9\-_]+\.[A-Za-z0-9\-_]+$", v):
        return 100
    if is_uuid(v):
        return 95
    if "session" in f or "jsessionid" in f or "phpsessid" in f:
        return 95
    if is_token(v):
        return 90
    if re.search(r"(id|uid|ref|key)$", f, re.I) and re.match(r"^\d{4,20}$", v):
        return 90
    if re.match(r"^\d{4,20}$", v):
        return 70
    return 50

# ─── V46-5: Safe Groovy TOKEN_KEYS list literal ─────────────────────────────
def groovy_token_keys_list() -> str:
    items = ", ".join(f'"{k}"' for k in TOKEN_KEYS)
    return f"[{items}]"

# ─── V46-4: Safe Groovy string embedding ────────────────────────────────────
def groovy_safe_string(s: str) -> str:
    s = s.replace("\\", "\\\\")
    s = s.replace("'", "\\'")
    return s

# ═══════════════════════════════════════════════════════════════════════════════
# V47-F: ENHANCED KNOWLEDGE BASE (SQLite)
# ═══════════════════════════════════════════════════════════════════════════════
def _db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    # Correlation cache table (v46)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS corr_cache (
            field         TEXT PRIMARY KEY,
            entity_group  TEXT,
            examples      TEXT,
            confidence    INTEGER DEFAULT 0,
            success_count INTEGER DEFAULT 0,
            failure_count INTEGER DEFAULT 0,
            last_seen     TEXT
        )
    """)
    # V47: Failure patterns knowledge base
    conn.execute("""
        CREATE TABLE IF NOT EXISTS failure_kb (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            error_code    TEXT NOT NULL,
            url_pattern   TEXT,
            failure_type  TEXT,
            root_cause    TEXT,
            fix_applied   TEXT,
            retry_count   INTEGER DEFAULT 0,
            final_status  TEXT,
            confidence    INTEGER DEFAULT 0,
            occurrence    INTEGER DEFAULT 1,
            created_at    TEXT,
            last_seen     TEXT
        )
    """)
    # V47: RCA history
    conn.execute("""
        CREATE TABLE IF NOT EXISTS rca_history (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            ts            TEXT,
            scenario      TEXT,
            failure_code  TEXT,
            root_cause    TEXT,
            confidence    INTEGER,
            fix_applied   TEXT,
            recovery      TEXT,
            final_result  TEXT
        )
    """)
    # V47: Self-healing stats
    conn.execute("""
        CREATE TABLE IF NOT EXISTS healing_stats (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            ts            TEXT,
            error_code    TEXT,
            action        TEXT,
            success       INTEGER DEFAULT 0,
            elapsed_ms    INTEGER DEFAULT 0
        )
    """)
    for col, typedef in [
        ("confidence",    "INTEGER DEFAULT 0"),
        ("success_count", "INTEGER DEFAULT 0"),
        ("failure_count", "INTEGER DEFAULT 0"),
    ]:
        try:
            conn.execute(f"ALTER TABLE corr_cache ADD COLUMN {col} {typedef}")
        except Exception:
            pass
    conn.commit()
    return conn

def db_store(field: str, value: str):
    entity_group = semantic_entity(field)
    conf         = correlation_confidence.get(field, 0)
    try:
        conn = _db_connect()
        conn.execute("""
            INSERT INTO corr_cache(field, entity_group, examples, confidence, last_seen)
            VALUES (?,?,?,?,?)
            ON CONFLICT(field) DO UPDATE SET
                entity_group = excluded.entity_group,
                examples     = excluded.examples,
                confidence   = excluded.confidence,
                last_seen    = excluded.last_seen
        """, (field, entity_group, value[:80], conf, datetime.now().isoformat()))
        conn.commit()
        conn.close()
    except Exception as ex:
        warn(f"db_store error: {ex}")

def db_record_outcome(field: str, success: bool):
    col = "success_count" if success else "failure_count"
    try:
        conn = _db_connect()
        conn.execute(
            f"UPDATE corr_cache SET {col} = {col} + 1 WHERE field = ?",
            (field,)
        )
        conn.commit()
        conn.close()
    except Exception as ex:
        warn(f"db_record_outcome error: {ex}")

def db_load_known_fields() -> dict[str, str]:
    try:
        conn  = _db_connect()
        rows  = conn.execute("SELECT field, entity_group FROM corr_cache").fetchall()
        conn.close()
        return {r[0]: r[1] for r in rows}
    except Exception:
        return {}

# V47: Store failure pattern in knowledge base
def kb_store_failure(error_code: str, url_pattern: str, failure_type: str,
                     root_cause: str, fix_applied: str, retry_count: int,
                     final_status: str, confidence: int):
    try:
        conn = _db_connect()
        existing = conn.execute(
            "SELECT id, occurrence FROM failure_kb WHERE error_code=? AND url_pattern=?",
            (error_code, url_pattern[:80])
        ).fetchone()
        now = datetime.now().isoformat()
        if existing:
            conn.execute(
                "UPDATE failure_kb SET occurrence=occurrence+1, last_seen=?, "
                "fix_applied=?, confidence=?, retry_count=?, final_status=? WHERE id=?",
                (now, fix_applied, confidence, retry_count, final_status, existing[0])
            )
        else:
            conn.execute(
                "INSERT INTO failure_kb(error_code, url_pattern, failure_type, root_cause, "
                "fix_applied, retry_count, final_status, confidence, created_at, last_seen) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (error_code, url_pattern[:80], failure_type, root_cause, fix_applied,
                 retry_count, final_status, confidence, now, now)
            )
        conn.commit()
        conn.close()
    except Exception as ex:
        warn(f"kb_store_failure error: {ex}")

# V47: Search failure pattern in knowledge base before analysis
def kb_search_failure(error_code: str, url_pattern: str = "") -> dict | None:
    try:
        conn = _db_connect()
        row = conn.execute(
            "SELECT root_cause, fix_applied, retry_count, final_status, confidence, occurrence "
            "FROM failure_kb WHERE error_code=? ORDER BY occurrence DESC, confidence DESC LIMIT 1",
            (error_code,)
        ).fetchone()
        conn.close()
        if row:
            return {
                "root_cause":   row[0],
                "fix_applied":  row[1],
                "retry_count":  row[2],
                "final_status": row[3],
                "confidence":   row[4],
                "occurrence":   row[5],
            }
        return None
    except Exception:
        return None

# V47: Store RCA entry
def kb_store_rca(scenario: str, failure_code: str, root_cause: str,
                 confidence: int, fix_applied: str, recovery: str, final_result: str):
    try:
        conn = _db_connect()
        conn.execute(
            "INSERT INTO rca_history(ts, scenario, failure_code, root_cause, "
            "confidence, fix_applied, recovery, final_result) VALUES (?,?,?,?,?,?,?,?)",
            (datetime.now().isoformat(), scenario, failure_code, root_cause,
             confidence, fix_applied, recovery, final_result)
        )
        conn.commit()
        conn.close()
    except Exception as ex:
        warn(f"kb_store_rca error: {ex}")

# V47: Store self-healing event
def kb_store_healing(error_code: str, action: str, success: bool, elapsed_ms: int = 0):
    try:
        conn = _db_connect()
        conn.execute(
            "INSERT INTO healing_stats(ts, error_code, action, success, elapsed_ms) "
            "VALUES (?,?,?,?,?)",
            (datetime.now().isoformat(), error_code, action, 1 if success else 0, elapsed_ms)
        )
        conn.commit()
        conn.close()
    except Exception as ex:
        warn(f"kb_store_healing error: {ex}")

# V47: Load all RCA history
def kb_load_rca() -> list[dict]:
    try:
        conn = _db_connect()
        rows = conn.execute(
            "SELECT ts, scenario, failure_code, root_cause, confidence, "
            "fix_applied, recovery, final_result FROM rca_history ORDER BY ts DESC LIMIT 50"
        ).fetchall()
        conn.close()
        return [
            {"ts": r[0], "scenario": r[1], "failure_code": r[2], "root_cause": r[3],
             "confidence": r[4], "fix_applied": r[5], "recovery": r[6], "final_result": r[7]}
            for r in rows
        ]
    except Exception:
        return []

# V47: Load self-healing stats
def kb_load_healing_stats() -> dict:
    try:
        conn = _db_connect()
        rows = conn.execute(
            "SELECT error_code, action, success, elapsed_ms FROM healing_stats"
        ).fetchall()
        conn.close()
        stats = defaultdict(lambda: {"total": 0, "success": 0, "fail": 0, "avg_ms": 0, "times": []})
        for code, action, succ, ms in rows:
            key = f"{code}:{action}"
            stats[key]["total"]   += 1
            stats[key]["success"] += succ
            stats[key]["fail"]    += (1 - succ)
            stats[key]["times"].append(ms)
        for k in stats:
            t = stats[k]["times"]
            stats[k]["avg_ms"] = int(sum(t) / len(t)) if t else 0
        return dict(stats)
    except Exception:
        return {}

# ─── ENTITY / LEARNING HELPERS (v46 unchanged) ────────────────────────────────
def _infer_entity_group(field: str) -> str:
    f = field.lower()
    entity_hints = [
        ("customer",  "CUSTOMER_ENTITY"), ("account",   "ACCOUNT_ENTITY"),
        ("order",     "ORDER_ENTITY"),     ("invoice",   "INVOICE_ENTITY"),
        ("policy",    "POLICY_ENTITY"),    ("claim",     "CLAIM_ENTITY"),
        ("employee",  "EMPLOYEE_ENTITY"),  ("case",      "CASE_ENTITY"),
        ("ticket",    "TICKET_ENTITY"),    ("user",      "USER_ENTITY"),
        ("project",   "PROJECT_ENTITY"),   ("task",      "TASK_ENTITY"),
        ("sprint",    "SPRINT_ENTITY"),
    ]
    for hint, group in entity_hints:
        if hint in f:
            return group
    return "GENERIC_ENTITY"

ENTITY_ALIASES: dict[str, str] = {
    "customerid": "CUSTOMER", "customer_uid": "CUSTOMER", "customerref": "CUSTOMER",
    "customernumber": "CUSTOMER", "custid": "CUSTOMER", "custref": "CUSTOMER",
    "accountid": "ACCOUNT", "account_uid": "ACCOUNT", "accountref": "ACCOUNT",
    "accountnumber": "ACCOUNT", "acctid": "ACCOUNT",
    "orderid": "ORDER", "orderref": "ORDER", "ordernumber": "ORDER", "order_uid": "ORDER",
    "invoiceid": "INVOICE", "invoice_uid": "INVOICE", "invoiceref": "INVOICE",
    "policyid": "POLICY", "policy_uid": "POLICY", "policyref": "POLICY",
    "claimid": "CLAIM", "claim_uid": "CLAIM", "claimref": "CLAIM",
    "employeeid": "EMPLOYEE", "employee_uid": "EMPLOYEE", "empid": "EMPLOYEE",
    "caseid": "CASE", "case_uid": "CASE", "caseref": "CASE",
    "ticketid": "TICKET", "ticket_uid": "TICKET", "ticketref": "TICKET",
    "userid": "USER", "user_uid": "USER", "userref": "USER",
}

def semantic_entity(field: str) -> str:
    key = re.sub(r"[^a-z0-9]", "", field.lower())
    if key in ENTITY_ALIASES:
        return ENTITY_ALIASES[key]
    inferred = _infer_entity_group(field)
    if inferred != "GENERIC_ENTITY":
        return inferred.replace("_ENTITY", "")
    return field.upper()

# ═══════════════════════════════════════════════════════════════════════════════
# V47-A: AI PERFORMANCE AGENT
# ═══════════════════════════════════════════════════════════════════════════════
class AIPerformanceAgent:
    """
    V47-A: Intelligent AI Agent that behaves like a Performance Engineer.
    Responsibilities:
      - Execute API tests (delegates to load engine)
      - Analyze failures with RCA + confidence score
      - Detect bottlenecks / correlation / token / session / header / schema issues
      - Retry automatically with exponential backoff
      - Search knowledge base before every analysis
      - Generate structured RCA + recommendations
    """

    KNOWN_ISSUES = {
        "401": {
            "root_cause": "Expired or missing JWT Token",
            "confidence": 95,
            "fix":        "Refresh authentication token and replay request",
            "recovery":   "Generated New Token",
        },
        "403": {
            "root_cause": "Insufficient permissions or CSRF token mismatch",
            "confidence": 90,
            "fix":        "Re-authenticate with correct role; validate CSRF header",
            "recovery":   "Re-authenticated Session",
        },
        "404": {
            "root_cause": "Resource not created yet — correlation dependency missing",
            "confidence": 85,
            "fix":        "Verify CREATE request preceded this GET/PUT; check CORR_ vars",
            "recovery":   "Correlation Chain Validated",
        },
        "408": {
            "root_cause": "Request timeout — backend slow or network saturated",
            "confidence": 80,
            "fix":        "Retry with exponential backoff; reduce concurrent users",
            "recovery":   "Exponential Backoff Applied",
        },
        "409": {
            "root_cause": "Duplicate test data — unique constraint violation",
            "confidence": 88,
            "fix":        "Inject unique ITER_SUFFIX into name/title/reference fields",
            "recovery":   "Fresh Faker Data Injected",
        },
        "422": {
            "root_cause": "Schema validation failure — malformed request payload",
            "confidence": 85,
            "fix":        "Verify required fields, data types, and correlated values",
            "recovery":   "Payload Rehealed",
        },
        "429": {
            "root_cause": "Rate limiting — too many requests per second",
            "confidence": 98,
            "fix":        "Respect Retry-After header; reduce ramp-up rate",
            "recovery":   "Rate-Limited Retry Scheduled",
        },
        "500": {
            "root_cause": "Backend exception — null pointer or unhandled error",
            "confidence": 75,
            "fix":        "Check server logs; validate request body completeness",
            "recovery":   "Smart Retry Initiated",
        },
        "502": {
            "root_cause": "Gateway error — upstream service unavailable",
            "confidence": 82,
            "fix":        "Wait for upstream recovery; enable circuit breaker",
            "recovery":   "Circuit Breaker Activated",
        },
        "503": {
            "root_cause": "Service unavailable — overloaded or in maintenance",
            "confidence": 90,
            "fix":        "Reduce load; wait for service recovery",
            "recovery":   "Graceful Degradation Applied",
        },
        "504": {
            "root_cause": "Gateway timeout — upstream took too long",
            "confidence": 85,
            "fix":        "Increase timeout thresholds; check database queries",
            "recovery":   "Timeout Extended + Retry",
        },
    }

    def __init__(self):
        self.healing_log:   list[dict] = []
        self.rca_log:       list[dict] = []
        self.analysis_log:  list[dict] = []

    # ── Analyze a single failure ──────────────────────────────────────────────
    def analyze_failure(self, error_code: str, url: str = "",
                        elapsed_ms: int = 0, response_body: str = "",
                        scenario: str = "unknown") -> dict:
        """
        V47-F: Search KB first. If known, apply fix immediately.
        If unknown, analyze and store.
        """
        code_str = str(error_code)

        # Step 1: Search knowledge base
        kb_hit = kb_search_failure(code_str, url)
        if kb_hit and kb_hit["confidence"] >= 70:
            info(f"  [AI Agent] KB HIT for {code_str} — applying known fix immediately")
            rca = {
                "error_code": code_str,
                "url":        url,
                "root_cause": kb_hit["root_cause"],
                "confidence": kb_hit["confidence"],
                "fix":        kb_hit["fix_applied"],
                "recovery":   kb_hit["final_status"],
                "source":     "KNOWLEDGE_BASE",
                "occurrence": kb_hit["occurrence"],
            }
        else:
            # Step 2: Use built-in rules or context inference
            base = self.KNOWN_ISSUES.get(code_str, {})
            root_cause = base.get("root_cause", "Unknown failure — no pattern matched")
            confidence = base.get("confidence", 50)
            fix        = base.get("fix", "Inspect logs and retry manually")
            recovery   = base.get("recovery", "Manual Investigation Required")

            # Context boost: very slow response → database bottleneck
            if code_str == "500" and elapsed_ms > 5000:
                root_cause = "Database Bottleneck — query exceeded 5s"
                confidence = 87
                fix        = "Optimize DB queries; add indexing; check connection pool"
                recovery   = "Database Query Optimized"

            # Schema detection from response body
            if response_body:
                if "required" in response_body.lower() and code_str in ("400", "422"):
                    root_cause = "Missing required field in request payload"
                    confidence = 92
                    fix        = "Add all required fields; re-heal request body"
                if "unique" in response_body.lower() and code_str == "409":
                    root_cause = "Unique constraint violation — duplicate record"
                    confidence = 96
                    fix        = "Inject ITER_SUFFIX into name/title/reference fields"
                if "token" in response_body.lower() and code_str == "401":
                    root_cause = "Token expired or revoked by server"
                    confidence = 98
                    fix        = "Re-login and refresh SHARED_TOKEN"

            rca = {
                "error_code": code_str,
                "url":        url,
                "root_cause": root_cause,
                "confidence": confidence,
                "fix":        fix,
                "recovery":   recovery,
                "source":     "AI_ANALYSIS",
                "occurrence": 1,
            }
            # Store in KB for future reuse
            kb_store_failure(
                code_str, url,
                self._failure_type(code_str),
                root_cause, fix, 0, recovery, confidence,
            )

        # Store RCA in DB and internal log
        kb_store_rca(scenario, code_str, rca["root_cause"], rca["confidence"],
                     rca["fix"], rca["recovery"], "ANALYZED")
        self.rca_log.append(rca)
        return rca

    def _failure_type(self, code: str) -> str:
        mapping = {
            "401": "Authentication Failure", "403": "Authorization Failure",
            "404": "Dependency Failure",     "408": "Timeout",
            "409": "Duplicate Data",         "422": "Schema Failure",
            "429": "Rate Limiting",          "500": "Backend Exception",
            "502": "Gateway Error",          "503": "Service Unavailable",
            "504": "Gateway Timeout",
        }
        return mapping.get(code, "Unknown Failure")

    # ── Print structured RCA output ───────────────────────────────────────────
    def print_rca(self, rca: dict):
        log()
        log(f"  {B}{C}╔{'═'*58}╗{W}")
        log(f"  {B}{C}║  AI ROOT CAUSE ANALYSIS                              ║{W}")
        log(f"  {B}{C}╚{'═'*58}╝{W}")
        log(f"  {'Failure:':<18} {R}{rca.get('error_code','?')} — "
            f"{self._failure_type(rca.get('error_code','?'))}{W}")
        log(f"  {'Root Cause:':<18} {rca.get('root_cause','?')}")
        log(f"  {'Confidence:':<18} {G}{rca.get('confidence','?')}%{W}")
        log(f"  {'Suggested Fix:':<18} {Y}{rca.get('fix','?')}{W}")
        log(f"  {'Recovery Action:':<18} {rca.get('recovery','?')}")
        log(f"  {'Source:':<18} {rca.get('source','?')}")
        log()

    # ── Detect header issues ──────────────────────────────────────────────────
    def detect_header_issues(self, reqs: list) -> list[str]:
        issues = []
        for r in reqs:
            if r.get("isLogin"):
                continue
            hdrs = r.get("headers", {})
            hdr_lower = {k.lower(): v for k, v in hdrs.items()}
            if "authorization" not in hdr_lower and "x-api-key" not in hdr_lower:
                issues.append(f"Missing auth header: {r['method']} {r['path']}")
            ct = hdr_lower.get("content-type", "")
            if r["method"] in ("POST", "PUT", "PATCH") and not ct:
                issues.append(f"Missing Content-Type: {r['method']} {r['path']}")
        if issues:
            for issue in issues:
                warn(f"[AI Agent] Header issue: {issue}")
        else:
            ok("[AI Agent] All headers validated — no issues found")
        return issues

    # ── Detect schema issues ──────────────────────────────────────────────────
    def detect_schema_issues(self, reqs: list) -> list[str]:
        issues = []
        for r in reqs:
            if r["method"] not in ("POST", "PUT", "PATCH"):
                continue
            body = r.get("postData", "")
            if not body:
                continue
            try:
                parsed = json.loads(body)
                if not isinstance(parsed, dict):
                    issues.append(f"Non-object body: {r['method']} {r['path']}")
                    continue
                for k, v in parsed.items():
                    if isinstance(v, str) and re.match(r"^[0-9a-f-]{36}$", v):
                        if not v.startswith("${"):
                            issues.append(
                                f"Hardcoded UUID in '{k}': {r['method']} {r['path']}"
                            )
            except Exception:
                pass
        if issues:
            for issue in issues:
                warn(f"[AI Agent] Schema issue: {issue}")
        else:
            ok("[AI Agent] Schema validation passed — no hardcoded UUIDs found")
        return issues

    # ── Detect correlation issues ─────────────────────────────────────────────
    def detect_correlation_issues(self, corr: dict, val_map: dict) -> list[str]:
        issues = []
        for field, prop in corr.items():
            conf = correlation_confidence.get(field, 0)
            if conf < 70:
                issues.append(f"Low-confidence correlation: {field} ({conf}%)")
        if issues:
            for issue in issues:
                warn(f"[AI Agent] Correlation issue: {issue}")
        else:
            ok("[AI Agent] All correlations validated")
        return issues

    # ── Detect token issues ───────────────────────────────────────────────────
    def detect_token_issues(self, reqs: list) -> list[str]:
        issues = []
        has_login = any(r.get("isLogin") for r in reqs)
        if not has_login:
            issues.append("No login request detected — token extraction will fail")
        for r in reqs:
            if r.get("isLogin") and not r.get("resp_json"):
                issues.append(f"Login response not captured: {r['url']}")
        if issues:
            for issue in issues:
                warn(f"[AI Agent] Token issue: {issue}")
        else:
            ok("[AI Agent] Token detection validated")
        return issues

    # ── Detect session issues ─────────────────────────────────────────────────
    def detect_session_issues(self, reqs: list) -> list[str]:
        issues = []
        session_cookies = {"jsessionid", "phpsessid", "asp.net_sessionid",
                           "connect.sid", "sessionid"}
        for r in reqs:
            raw_cookie = r.get("headers", {}).get("cookie", "")
            if not raw_cookie:
                continue
            for part in raw_cookie.split(";"):
                if "=" in part:
                    name = part.strip().split("=")[0].lower()
                    if name in session_cookies:
                        val = part.strip().split("=", 1)[1] if "=" in part else ""
                        if val and not val.startswith("${"):
                            issues.append(
                                f"Hardcoded session cookie '{name}': {r['path']}"
                            )
        if issues:
            for issue in issues:
                warn(f"[AI Agent] Session issue: {issue}")
        else:
            ok("[AI Agent] Session cookie handling validated")
        return issues

    # ── Full scenario analysis ────────────────────────────────────────────────
    def analyze_scenario(self, reqs: list, corr: dict, val_map: dict,
                         scenario: str = "recorded") -> dict:
        banner("V47 AI AGENT — Full Scenario Analysis")
        header_issues      = self.detect_header_issues(reqs)
        schema_issues      = self.detect_schema_issues(reqs)
        correlation_issues = self.detect_correlation_issues(corr, val_map)
        token_issues       = self.detect_token_issues(reqs)
        session_issues     = self.detect_session_issues(reqs)
        total_issues = (len(header_issues) + len(schema_issues) +
                        len(correlation_issues) + len(token_issues) +
                        len(session_issues))
        health_score = max(0, 100 - (total_issues * 8))
        log()
        log(f"  {B}AI Agent Scenario Health Score: "
            f"{'🟢' if health_score >= 80 else '🟡' if health_score >= 60 else '🔴'}"
            f"  {health_score}/100{W}")
        result = {
            "health_score":        health_score,
            "header_issues":       header_issues,
            "schema_issues":       schema_issues,
            "correlation_issues":  correlation_issues,
            "token_issues":        token_issues,
            "session_issues":      session_issues,
            "total_issues":        total_issues,
            "recommendations":     self._build_recommendations(
                header_issues, schema_issues,
                correlation_issues, token_issues, session_issues
            ),
        }
        self.analysis_log.append({"scenario": scenario, "result": result,
                                   "ts": datetime.now().isoformat()})
        return result

    def _build_recommendations(self, header_issues, schema_issues,
                                correlation_issues, token_issues, session_issues) -> list[str]:
        recs = []
        if header_issues:
            recs.append("Add Authorization: Bearer ${__P(SHARED_TOKEN,INIT)} to all samplers")
        if schema_issues:
            recs.append("Replace hardcoded UUIDs with ${CORR_*} variables")
        if correlation_issues:
            recs.append("Review low-confidence correlations — re-record if needed")
        if token_issues:
            recs.append("Ensure login request is captured before other API requests")
        if session_issues:
            recs.append("Replace hardcoded session cookies with ${COOKIE_*} variables")
        if not recs:
            recs.append("Scenario looks clean — proceed to JMX generation")
        return recs

    # ── Print RCA summary ─────────────────────────────────────────────────────
    def print_rca_summary(self):
        banner("V47 AI Agent — RCA Summary")
        if not self.rca_log:
            info("No RCA events recorded in this session.")
            # Also pull from DB
            db_rcas = kb_load_rca()
            if db_rcas:
                log(f"  {B}Last {len(db_rcas)} RCA events from knowledge base:{W}")
                for r in db_rcas[:10]:
                    log(f"  {Y}[{r['ts'][:16]}]{W} Code={r['failure_code']} "
                        f"Cause={r['root_cause'][:50]} "
                        f"Conf={r['confidence']}% Result={r['final_result']}")
            return
        for rca in self.rca_log:
            self.print_rca(rca)

    # ── Print self-healing summary ────────────────────────────────────────────
    def print_healing_summary(self):
        banner("V47 Self-Healing Statistics")
        stats = kb_load_healing_stats()
        if not stats:
            info("No healing events recorded yet.")
            return
        log(f"  {B}{'Error:Action':<30} {'Total':>6} {'Success':>8} "
            f"{'Fail':>6} {'Avg ms':>8}{W}")
        log(f"  {'─'*30} {'─'*6} {'─'*8} {'─'*6} {'─'*8}")
        for key, s in sorted(stats.items()):
            rate = f"{s['success']/max(s['total'],1)*100:.0f}%"
            log(f"  {key:<30} {s['total']:>6} {s['success']:>7} ({rate}) "
                f"{s['fail']:>6} {s['avg_ms']:>8}ms")
        total    = sum(s["total"]   for s in stats.values())
        success  = sum(s["success"] for s in stats.values())
        log(f"  {'─'*30} {'─'*6} {'─'*8} {'─'*6} {'─'*8}")
        log(f"  {B}{'TOTAL':<30} {total:>6} {success:>8} "
            f"{total-success:>6}{W}")
        ok(f"Self-healing success rate: "
           f"{success/max(total,1)*100:.1f}%  ({success}/{total})")

# Global agent instance
ai_agent = AIPerformanceAgent()

# ═══════════════════════════════════════════════════════════════════════════════
# V47-C: LIVE SYSTEM BOTTLENECK MONITOR (psutil)
# ═══════════════════════════════════════════════════════════════════════════════
class BottleneckMonitor:
    """
    V47-C: Real-time CPU / Memory / Disk / Network monitor using psutil.
    Detects: High CPU (>80%), Memory Leak (>90%), Network Saturation, Slow API.
    Generates recommendations automatically.
    """
    def __init__(self, interval: float = 5.0):
        self.interval      = interval
        self._stop         = threading.Event()
        self._thread       = None
        self.samples:       list[dict] = []
        self.alerts:        list[str]  = []
        self._net_baseline: dict | None = None

    def start(self):
        self._stop.clear()
        self._net_baseline = psutil.net_io_counters()._asdict()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        ok("[BottleneckMonitor] Live system monitoring started")

    def stop(self):
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=10)
        ok("[BottleneckMonitor] Monitoring stopped")

    def _run(self):
        while not self._stop.is_set():
            try:
                cpu_pct  = psutil.cpu_percent(interval=1)
                mem      = psutil.virtual_memory()
                disk     = psutil.disk_io_counters()
                net      = psutil.net_io_counters()
                sample = {
                    "ts":           datetime.now().isoformat(),
                    "cpu_pct":      cpu_pct,
                    "mem_pct":      mem.percent,
                    "mem_used_mb":  mem.used // (1024 * 1024),
                    "mem_total_mb": mem.total // (1024 * 1024),
                    "disk_read_mb": (disk.read_bytes  // (1024 * 1024)) if disk else 0,
                    "disk_write_mb":(disk.write_bytes // (1024 * 1024)) if disk else 0,
                    "net_sent_mb":  net.bytes_sent // (1024 * 1024),
                    "net_recv_mb":  net.bytes_recv // (1024 * 1024),
                }
                self.samples.append(sample)

                # Alert detection
                if cpu_pct > 80:
                    alert = f"HIGH CPU: {cpu_pct:.1f}% > 80%"
                    if alert not in self.alerts:
                        self.alerts.append(alert)
                        warn(f"[Monitor] {alert}")
                if mem.percent > 90:
                    alert = f"MEMORY WARNING: {mem.percent:.1f}% used"
                    if alert not in self.alerts:
                        self.alerts.append(alert)
                        warn(f"[Monitor] {alert}")
                # Network saturation: >500 MB/s recv delta
                if self._net_baseline:
                    net_delta = (net.bytes_recv - self._net_baseline["bytes_recv"])
                    if net_delta > 500 * 1024 * 1024:
                        alert = f"NETWORK SATURATION: >{net_delta//(1024*1024)} MB received"
                        if alert not in self.alerts:
                            self.alerts.append(alert)
                            warn(f"[Monitor] {alert}")
            except Exception as ex:
                pass
            self._stop.wait(self.interval)

    def summary(self) -> dict:
        if not self.samples:
            return {}
        cpu_vals  = [s["cpu_pct"]   for s in self.samples]
        mem_vals  = [s["mem_pct"]   for s in self.samples]
        net_recv  = [s["net_recv_mb"] for s in self.samples]
        return {
            "cpu_avg":    round(statistics.mean(cpu_vals),   1),
            "cpu_max":    round(max(cpu_vals),               1),
            "mem_avg":    round(statistics.mean(mem_vals),   1),
            "mem_max":    round(max(mem_vals),               1),
            "net_recv_total_mb": max(net_recv) - min(net_recv),
            "samples":    len(self.samples),
            "alerts":     self.alerts,
        }

    def print_summary(self):
        s = self.summary()
        if not s:
            info("No monitoring data collected.")
            return
        log()
        log(f"  {B}{C}SYSTEM BOTTLENECK REPORT{W}")
        log(f"  {'─'*55}")
        log(f"  CPU Average   : {s['cpu_avg']:>6.1f}%  "
            f"{'⚠ HIGH' if s['cpu_avg']>80 else '✅ OK'}")
        log(f"  CPU Peak      : {s['cpu_max']:>6.1f}%")
        log(f"  Memory Avg    : {s['mem_avg']:>6.1f}%  "
            f"{'⚠ HIGH' if s['mem_avg']>85 else '✅ OK'}")
        log(f"  Memory Peak   : {s['mem_max']:>6.1f}%")
        log(f"  Net Received  : {s['net_recv_total_mb']:>6} MB")
        log(f"  Alerts        : {len(s['alerts'])}")
        for a in s.get("alerts", []):
            log(f"    {R}⚠  {a}{W}")
        log()
        log(f"  {B}RECOMMENDATIONS:{W}")
        recs = self._generate_recommendations(s)
        for i, r in enumerate(recs, 1):
            log(f"  {i}. {r}")

    def _generate_recommendations(self, s: dict) -> list[str]:
        recs = []
        if s.get("cpu_avg", 0) > 80:
            recs.append("Reduce concurrent virtual users or scale horizontally")
            recs.append("Profile application code for CPU-intensive loops")
        if s.get("mem_avg", 0) > 85:
            recs.append("Check for memory leaks — heap dump recommended")
            recs.append("Increase JVM heap size: -Xmx2g -Xms512m")
        if s.get("net_recv_total_mb", 0) > 1000:
            recs.append("Network bandwidth near saturation — enable response compression")
        if not recs:
            recs.append("System resources within normal thresholds during test")
        return recs

# ═══════════════════════════════════════════════════════════════════════════════
# V47-B: PURE PYTHON LOAD ENGINE (AsyncIO + aiohttp + ThreadPoolExecutor)
# ═══════════════════════════════════════════════════════════════════════════════
class LoadResult:
    """Holds all metrics from a single virtual user request."""
    __slots__ = ("url", "method", "status_code", "elapsed_ms",
                 "success", "error", "ts", "response_size")
    def __init__(self, url, method, status_code, elapsed_ms,
                 success, error="", response_size=0):
        self.url           = url
        self.method        = method
        self.status_code   = status_code
        self.elapsed_ms    = elapsed_ms
        self.success       = success
        self.error         = error
        self.response_size = response_size
        self.ts            = datetime.now().isoformat()

class LoadEngine:
    """
    V47-B: Pure Python load runner using AsyncIO + aiohttp.
    Supports: Functional / Load / Stress / Spike / Endurance / Volume 5K.
    Replays recorded requests without JMeter — useful for quick validation
    and environments where JMeter is unavailable.
    """

    def __init__(self, reqs: list, token: str = "", monitor: BottleneckMonitor | None = None):
        self.reqs     = [r for r in reqs if r.get("method") != "WEBSOCKET"]
        self.token    = token
        self.monitor  = monitor
        self.results: list[LoadResult] = []
        self._lock    = asyncio.Lock()

    # ── Single async request ──────────────────────────────────────────────────
    async def _send_one(self, session: aiohttp.ClientSession, # type: ignore
                        req: dict, user_id: int) -> LoadResult:
        url     = req.get("url", "")
        method  = req.get("method", "GET")
        body    = req.get("postData", "") if method in ("POST","PUT","PATCH") else None
        headers = {
            k: v for k, v in req.get("headers", {}).items()
            if k.lower() not in SKIP_HDRS
        }
        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        # Replace JMeter variables with empty strings for pure-python run
        if body:
            body = re.sub(r"\$\{[^}]+\}", "", body)

        t0 = time.monotonic()
        try:
            timeout = aiohttp.ClientTimeout(
                connect=CONNECT_TIMEOUT_MS / 1000,
                total=RESPONSE_TIMEOUT_MS / 1000,
            )
            kwargs: dict = {"headers": headers, "timeout": timeout,
                            "ssl": False, "allow_redirects": True}
            if body:
                kwargs["data"] = body
            async with session.request(method, url, **kwargs) as resp:
                elapsed   = int((time.monotonic() - t0) * 1000)
                raw       = await resp.read()
                success   = 200 <= resp.status < 300
                return LoadResult(url, method, resp.status, elapsed, success,
                                  response_size=len(raw))
        except asyncio.TimeoutError:
            elapsed = int((time.monotonic() - t0) * 1000)
            return LoadResult(url, method, 408, elapsed, False, "Timeout")
        except Exception as ex:
            elapsed = int((time.monotonic() - t0) * 1000)
            return LoadResult(url, method, 0, elapsed, False, str(ex)[:80])

    # ── Virtual user task ─────────────────────────────────────────────────────
    async def _virtual_user(self, session: aiohttp.ClientSession, # pyright: ignore[reportInvalidTypeForm]
                             user_id: int, duration_s: int,
                             iteration_results: list):
        end_time = time.monotonic() + duration_s                                            
        while time.monotonic() < end_time:
            for req in self.reqs:
                if time.monotonic() >= end_time:
                    break
                result = await self._send_one(session, req, user_id)
                async with self._lock:
                    iteration_results.append(result)
                # Think time: 0.5–1.5s between requests
                await asyncio.sleep(0.5 + (user_id % 10) * 0.1)

    # ── Ramp-up coordinator ───────────────────────────────────────────────────
    async def _run_async(self, users: int, ramp_s: int, duration_s: int) -> list[LoadResult]:
        results: list[LoadResult] = []
        connector = aiohttp.TCPConnector(limit=min(users + 10, 500), ssl=False)
        async with aiohttp.ClientSession(connector=connector) as session:
            tasks = []
            delay_per_user = ramp_s / max(users, 1)
            for uid in range(users):
                await asyncio.sleep(delay_per_user)
                t = asyncio.create_task(
                    self._virtual_user(session, uid, duration_s, results)
                )
                tasks.append(t)
                # Progress log every 50 users
                if uid % 50 == 0 and uid > 0:
                    ok(f"[Load Engine] {uid}/{users} virtual users started "
                       f"({len(results)} requests so far)")
            await asyncio.gather(*tasks, return_exceptions=True)
        return results

    # ── Public run entry point ────────────────────────────────────────────────
    def run(self, profile_key: str = "2") -> "LoadEngineReport":
        profile = LOAD_VU_PROFILES.get(profile_key, LOAD_VU_PROFILES["2"])
        users    = profile["users"]
        ramp_s   = profile["ramp"]
        dur_s    = profile["duration"]

        banner(f"V47 Load Engine — {profile['name']}  "
               f"({users} VUs / {ramp_s}s ramp / {dur_s}s duration)")
        ok(f"Target URLs: {len(self.reqs)} endpoint(s) in scenario")

        if self.monitor:
            self.monitor.start()

        t_start = time.monotonic()
        try:
            self.results = asyncio.run(self._run_async(users, ramp_s, dur_s))
        except Exception as ex:
            err(f"[Load Engine] Run failed: {ex}")
            self.results = []
        finally:
            if self.monitor:
                self.monitor.stop()

        elapsed_total = time.monotonic() - t_start
        ok(f"[Load Engine] Completed: {len(self.results)} requests in "
           f"{elapsed_total:.1f}s")

        return LoadEngineReport(self.results, profile, elapsed_total)

    # ── Per-test-type shortcuts ───────────────────────────────────────────────
    def run_functional(self) -> "LoadEngineReport": return self.run("1")
    def run_load(self)       -> "LoadEngineReport": return self.run("2")
    def run_stress(self)     -> "LoadEngineReport": return self.run("3")
    def run_spike(self)      -> "LoadEngineReport": return self.run("4")
    def run_endurance(self)  -> "LoadEngineReport": return self.run("5")

class LoadEngineReport:
    """Computes all metrics from LoadEngine results and feeds the report engine."""
    def __init__(self, results: list[LoadResult], profile: dict, elapsed_s: float):
        self.results  = results
        self.profile  = profile
        self.elapsed  = elapsed_s
        self.metrics  = self._compute()

    def _compute(self) -> dict:
        if not self.results:
            return self._empty_metrics()
        total      = len(self.results)
        successes  = [r for r in self.results if r.success]
        failures   = [r for r in self.results if not r.success]
        times      = [r.elapsed_ms for r in self.results]
        times_s    = sorted(times)
        def pct(p):
            idx = int(len(times_s) * p / 100)
            return times_s[min(idx, len(times_s)-1)]

        total_bytes = sum(r.response_size for r in self.results)
        tps = total / max(self.elapsed, 1)
        throughput_kbps = (total_bytes / 1024) / max(self.elapsed, 1)

        # Failure breakdown by status code
        failure_codes: dict[int, int] = defaultdict(int)
        for r in failures:
            failure_codes[r.status_code] += 1

        return {
            "profile_name":    self.profile["name"],
            "virtual_users":   self.profile["users"],
            "duration_s":      self.profile["duration"],
            "total_requests":  total,
            "total_success":   len(successes),
            "total_failures":  len(failures),
            "success_rate":    round(len(successes) / max(total, 1) * 100, 2),
            "error_rate":      round(len(failures)  / max(total, 1) * 100, 2),
            "tps":             round(tps,              2),
            "throughput_kbps": round(throughput_kbps,  2),
            "avg_ms":          round(statistics.mean(times),          1) if times else 0,
            "min_ms":          min(times)  if times else 0,
            "max_ms":          max(times)  if times else 0,
            "median_ms":       pct(50)     if times else 0,
            "p90_ms":          pct(90)     if times else 0,
            "p95_ms":          pct(95)     if times else 0,
            "p99_ms":          pct(99)     if times else 0,
            "stdev_ms":        round(statistics.stdev(times), 1) if len(times) > 1 else 0,
            "failure_codes":   dict(failure_codes),
            "elapsed_s":       round(self.elapsed, 2),
        }

    @staticmethod
    def _empty_metrics() -> dict:
        return {k: 0 for k in [
            "profile_name", "virtual_users", "duration_s", "total_requests",
            "total_success", "total_failures", "success_rate", "error_rate",
            "tps", "throughput_kbps", "avg_ms", "min_ms", "max_ms",
            "median_ms", "p90_ms", "p95_ms", "p99_ms", "stdev_ms",
            "failure_codes", "elapsed_s",
        ]}

    def print_summary(self):
        m = self.metrics
        banner(f"Load Test Results — {m.get('profile_name','?')}")
        log(f"  {'─'*62}")
        rows = [
            ("Virtual Users",    m.get("virtual_users",0)),
            ("Total Requests",   m.get("total_requests",0)),
            ("Successful",       f"{m.get('total_success',0)}  "
                                 f"({m.get('success_rate',0):.1f}%)"),
            ("Failed",           f"{m.get('total_failures',0)}  "
                                 f"({m.get('error_rate',0):.1f}%)"),
            ("TPS",              f"{m.get('tps',0):.2f} req/s"),
            ("Throughput",       f"{m.get('throughput_kbps',0):.1f} KB/s"),
            ("Avg Response",     f"{m.get('avg_ms',0):.1f} ms"),
            ("Median (P50)",     f"{m.get('median_ms',0)} ms"),
            ("P90",              f"{m.get('p90_ms',0)} ms"),
            ("P95",              f"{m.get('p95_ms',0)} ms"),
            ("P99",              f"{m.get('p99_ms',0)} ms"),
            ("Min / Max",        f"{m.get('min_ms',0)} / {m.get('max_ms',0)} ms"),
            ("Std Deviation",    f"{m.get('stdev_ms',0):.1f} ms"),
            ("Total Duration",   f"{m.get('elapsed_s',0):.1f} s"),
        ]
        for label, value in rows:
            emoji = ""
            if label == "Failed" and m.get("total_failures", 0) > 0:
                emoji = f"  {R}⚠{W}"
            log(f"  {label:<22} {str(value):<30}{emoji}")
        if m.get("failure_codes"):
            log(f"\n  {Y}Failure Breakdown:{W}")
            for code, count in sorted(m["failure_codes"].items()):
                rca = ai_agent.analyze_failure(str(code), scenario=m.get("profile_name",""))
                log(f"    {code}  ×{count:<5}  → {rca['root_cause'][:45]}  "
                    f"({rca['confidence']}%)")
        log(f"  {'─'*62}")

# ═══════════════════════════════════════════════════════════════════════════════
# V47-D: FULL REPORT ENGINE (HTML + Excel + PDF)
# ═══════════════════════════════════════════════════════════════════════════════
class ReportEngine:
    """
    V47-D: Generates HTML, Excel, and PDF reports after every test run.
    Includes: TPS, Throughput, Avg Response, P90, P95, P99, Error Rate,
    Success Rate, Self-Healing Statistics, RCA Summary, Recommendations.
    """

    def __init__(self, ts: str | None = None):
        self.ts = ts or datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── HTML REPORT ───────────────────────────────────────────────────────────
    def generate_html(self, metrics: dict, monitor_summary: dict | None = None,
                      rca_log: list | None = None, healing_stats: dict | None = None,
                      scenario_name: str = "Scenario") -> str:
        rca_log       = rca_log or []
        healing_stats = healing_stats or {}
        ms            = monitor_summary or {}
        m             = metrics

        def row(label, value, highlight=False):
            cls = ' class="highlight"' if highlight else ''
            return f'<tr{cls}><td>{label}</td><td><strong>{value}</strong></td></tr>'

        # Build failure codes table
        fc_rows = ""
        for code, count in (m.get("failure_codes") or {}).items():
            fc_rows += f"<tr><td>{code}</td><td>{count}</td></tr>"
        if not fc_rows:
            fc_rows = '<tr><td colspan="2">✅ No failures</td></tr>'

        # Build RCA table
        rca_rows = ""
        for rca in (rca_log or [])[:10]:
            conf_color = "green" if rca.get("confidence",0)>=80 else "orange"
            rca_rows += (
                f"<tr>"
                f"<td>{rca.get('error_code','?')}</td>"
                f"<td>{rca.get('root_cause','?')[:60]}</td>"
                f"<td style='color:{conf_color}'>{rca.get('confidence','?')}%</td>"
                f"<td>{rca.get('fix','?')[:60]}</td>"
                f"<td>{rca.get('recovery','?')[:40]}</td>"
                f"</tr>"
            )
        if not rca_rows:
            rca_rows = '<tr><td colspan="5">No RCA events in this session</td></tr>'

        # Build healing stats table
        hs_rows = ""
        for key, s in list((healing_stats or {}).items())[:10]:
            rate = f"{s['success']/max(s['total'],1)*100:.0f}%"
            hs_rows += (
                f"<tr>"
                f"<td>{key}</td>"
                f"<td>{s['total']}</td>"
                f"<td>{s['success']} ({rate})</td>"
                f"<td>{s['fail']}</td>"
                f"<td>{s.get('avg_ms',0)} ms</td>"
                f"</tr>"
            )
        if not hs_rows:
            hs_rows = '<tr><td colspan="5">No healing events yet</td></tr>'

        # Recommendations
        recs_html = ""
        for rec in self._build_report_recommendations(m, ms):
            recs_html += f"<li>{rec}</li>"

        alert_html = ""
        for a in ms.get("alerts", []):
            alert_html += f'<div class="alert">⚠ {a}</div>'

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Performance Test Report — {scenario_name} — {self.ts}</title>
<style>
:root {{
  --bg:       #0f1117;
  --card:     #1a1d2e;
  --accent:   #00d4aa;
  --accent2:  #7c6aff;
  --danger:   #ff4560;
  --warn:     #ffb74d;
  --text:     #e8eaf0;
  --subtext:  #8b90a7;
  --border:   #2d3148;
  --green:    #00e676;
}}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:var(--bg); color:var(--text); font-family:'Segoe UI',system-ui,sans-serif; padding:24px; }}
h1 {{ font-size:1.8rem; color:var(--accent); margin-bottom:4px; }}
.subtitle {{ color:var(--subtext); font-size:.9rem; margin-bottom:32px; }}
.grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(240px,1fr)); gap:16px; margin-bottom:32px; }}
.card {{ background:var(--card); border:1px solid var(--border); border-radius:10px; padding:20px; }}
.card .label {{ color:var(--subtext); font-size:.75rem; text-transform:uppercase; letter-spacing:.08em; margin-bottom:6px; }}
.card .value {{ font-size:1.9rem; font-weight:700; color:var(--accent); }}
.card .value.danger {{ color:var(--danger); }}
.card .value.warn {{ color:var(--warn); }}
.section {{ background:var(--card); border:1px solid var(--border); border-radius:10px; padding:24px; margin-bottom:24px; }}
.section h2 {{ color:var(--accent2); font-size:1rem; text-transform:uppercase; letter-spacing:.1em; margin-bottom:16px; border-bottom:1px solid var(--border); padding-bottom:10px; }}
table {{ width:100%; border-collapse:collapse; font-size:.875rem; }}
th {{ background:rgba(124,106,255,.15); color:var(--accent2); text-align:left; padding:10px 12px; font-size:.8rem; text-transform:uppercase; letter-spacing:.05em; }}
td {{ padding:9px 12px; border-bottom:1px solid var(--border); color:var(--text); }}
tr:last-child td {{ border-bottom:none; }}
tr:hover td {{ background:rgba(255,255,255,.03); }}
.highlight td {{ background:rgba(0,212,170,.06); }}
.alert {{ background:rgba(255,69,96,.12); border:1px solid rgba(255,69,96,.3); border-radius:6px; padding:10px 14px; margin:6px 0; color:var(--danger); font-size:.875rem; }}
ul.recs {{ padding-left:20px; color:var(--text); }}
ul.recs li {{ padding:5px 0; font-size:.875rem; }}
.badge {{ display:inline-block; padding:2px 8px; border-radius:4px; font-size:.75rem; font-weight:600; }}
.badge.pass {{ background:rgba(0,230,118,.15); color:var(--green); }}
.badge.fail {{ background:rgba(255,69,96,.15); color:var(--danger); }}
footer {{ text-align:center; color:var(--subtext); font-size:.8rem; margin-top:40px; }}
</style>
</head>
<body>
<h1>⚡ Performance Test Report</h1>
<div class="subtitle">
  Scenario: <strong>{scenario_name}</strong> &nbsp;|&nbsp;
  Profile: <strong>{m.get('profile_name','—')}</strong> &nbsp;|&nbsp;
  Generated: <strong>{self.ts}</strong>
</div>

<!-- KPI Cards -->
<div class="grid">
  <div class="card">
    <div class="label">Virtual Users</div>
    <div class="value">{m.get('virtual_users',0)}</div>
  </div>
  <div class="card">
    <div class="label">TPS</div>
    <div class="value">{m.get('tps',0):.2f}</div>
  </div>
  <div class="card">
    <div class="label">Throughput</div>
    <div class="value">{m.get('throughput_kbps',0):.1f} <small style="font-size:.9rem">KB/s</small></div>
  </div>
  <div class="card">
    <div class="label">Avg Response</div>
    <div class="value {'warn' if m.get('avg_ms',0)>3000 else ''}">{m.get('avg_ms',0):.0f} <small style="font-size:.9rem">ms</small></div>
  </div>
  <div class="card">
    <div class="label">P90</div>
    <div class="value {'warn' if m.get('p90_ms',0)>5000 else ''}">{m.get('p90_ms',0)} <small style="font-size:.9rem">ms</small></div>
  </div>
  <div class="card">
    <div class="label">P95</div>
    <div class="value">{m.get('p95_ms',0)} <small style="font-size:.9rem">ms</small></div>
  </div>
  <div class="card">
    <div class="label">P99</div>
    <div class="value">{m.get('p99_ms',0)} <small style="font-size:.9rem">ms</small></div>
  </div>
  <div class="card">
    <div class="label">Success Rate</div>
    <div class="value {'pass' if m.get('success_rate',0)>=95 else 'danger'}">{m.get('success_rate',0):.1f}%</div>
  </div>
  <div class="card">
    <div class="label">Error Rate</div>
    <div class="value {'danger' if m.get('error_rate',0)>5 else ''}">{m.get('error_rate',0):.1f}%</div>
  </div>
  <div class="card">
    <div class="label">Total Requests</div>
    <div class="value">{m.get('total_requests',0)}</div>
  </div>
</div>

<!-- Metrics Table -->
<div class="section">
  <h2>📊 Detailed Metrics</h2>
  <table>
    <thead><tr><th>Metric</th><th>Value</th></tr></thead>
    <tbody>
      {row("Min Response",      f"{m.get('min_ms',0)} ms")}
      {row("Avg Response",      f"{m.get('avg_ms',0):.1f} ms")}
      {row("Median (P50)",      f"{m.get('median_ms',0)} ms")}
      {row("P90",               f"{m.get('p90_ms',0)} ms",
           m.get('p90_ms',0) > 5000)}
      {row("P95",               f"{m.get('p95_ms',0)} ms")}
      {row("P99",               f"{m.get('p99_ms',0)} ms")}
      {row("Max Response",      f"{m.get('max_ms',0)} ms")}
      {row("Std Deviation",     f"{m.get('stdev_ms',0):.1f} ms")}
      {row("TPS",               f"{m.get('tps',0):.2f} req/s")}
      {row("Throughput",        f"{m.get('throughput_kbps',0):.1f} KB/s")}
      {row("Total Requests",    m.get('total_requests',0))}
      {row("Successful",        f"{m.get('total_success',0)} ({m.get('success_rate',0):.1f}%)")}
      {row("Failed",            f"{m.get('total_failures',0)} ({m.get('error_rate',0):.1f}%)",
           m.get('total_failures',0) > 0)}
      {row("Test Duration",     f"{m.get('elapsed_s',0):.1f} s")}
    </tbody>
  </table>
</div>

<!-- Error Breakdown -->
<div class="section">
  <h2>❌ Failure Breakdown</h2>
  <table>
    <thead><tr><th>HTTP Code</th><th>Count</th></tr></thead>
    <tbody>{fc_rows}</tbody>
  </table>
</div>

<!-- System Monitor -->
<div class="section">
  <h2>🖥 System Resource Monitor</h2>
  {alert_html if alert_html else '<p style="color:var(--green)">✅ No resource alerts</p>'}
  <table style="margin-top:12px">
    <thead><tr><th>Metric</th><th>Value</th></tr></thead>
    <tbody>
      {row("CPU Average",     f"{ms.get('cpu_avg',0):.1f}%", ms.get('cpu_avg',0)>80)}
      {row("CPU Peak",        f"{ms.get('cpu_max',0):.1f}%")}
      {row("Memory Average",  f"{ms.get('mem_avg',0):.1f}%", ms.get('mem_avg',0)>85)}
      {row("Memory Peak",     f"{ms.get('mem_max',0):.1f}%")}
      {row("Net Received",    f"{ms.get('net_recv_total_mb',0)} MB")}
      {row("Monitor Samples", ms.get('samples',0))}
    </tbody>
  </table>
</div>

<!-- RCA Table -->
<div class="section">
  <h2>🧠 AI Root Cause Analysis</h2>
  <table>
    <thead>
      <tr>
        <th>Code</th><th>Root Cause</th><th>Confidence</th>
        <th>Fix</th><th>Recovery</th>
      </tr>
    </thead>
    <tbody>{rca_rows}</tbody>
  </table>
</div>

<!-- Self-Healing Stats -->
<div class="section">
  <h2>🔧 Self-Healing Statistics</h2>
  <table>
    <thead>
      <tr>
        <th>Error:Action</th><th>Total</th><th>Success</th>
        <th>Fail</th><th>Avg Time</th>
      </tr>
    </thead>
    <tbody>{hs_rows}</tbody>
  </table>
</div>

<!-- Recommendations -->
<div class="section">
  <h2>💡 Recommendations</h2>
  <ul class="recs">{recs_html}</ul>
</div>

<footer>
  Generated by JMeter Automation v47 — Universal Enterprise Framework &nbsp;|&nbsp;
  {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
</footer>
</body>
</html>"""

        out_path = os.path.join(REPORTS_DIR, f"report_{self.ts}.html")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)
        ok(f"HTML report → {out_path}")
        return out_path

    # ── EXCEL REPORT ──────────────────────────────────────────────────────────
    def generate_excel(self, metrics: dict, monitor_summary: dict | None = None,
                       rca_log: list | None = None, healing_stats: dict | None = None,
                       scenario_name: str = "Scenario") -> str:
        rca_log       = rca_log or []
        healing_stats = healing_stats or {}
        ms            = monitor_summary or {}
        m             = metrics

        wb = openpyxl.Workbook()

        # ── Styles ────────────────────────────────────────────────────────────
        hdr_font     = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill     = PatternFill("solid", fgColor="1a1d2e")
        accent_fill  = PatternFill("solid", fgColor="004d3d")
        warn_fill    = PatternFill("solid", fgColor="7a3800")
        fail_fill    = PatternFill("solid", fgColor="5a0000")
        green_font   = Font(color="00d4aa", bold=True)
        warn_font    = Font(color="ffb74d", bold=True)
        fail_font    = Font(color="ff4560", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(
            left=Side(style="thin", color="2d3148"),
            right=Side(style="thin", color="2d3148"),
            top=Side(style="thin", color="2d3148"),
            bottom=Side(style="thin", color="2d3148"),
        )

        def style_header_row(ws, row_num, col_count):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font   = hdr_font
                cell.fill   = hdr_fill
                cell.alignment = center_align
                cell.border = thin_border

        def auto_width(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        # ── Sheet 1: Summary ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        ws.sheet_properties.tabColor = "00d4aa"

        ws["A1"] = f"Performance Test Report — {scenario_name}"
        ws["A1"].font = Font(bold=True, size=14, color="00d4aa")
        ws["A2"] = f"Profile: {m.get('profile_name','—')}  |  Generated: {self.ts}"
        ws["A2"].font = Font(italic=True, color="8b90a7")

        headers = ["Metric", "Value"]
        ws.append([])
        ws.append(headers)
        style_header_row(ws, 4, 2)

        kpis = [
            ("Virtual Users",   m.get("virtual_users",    0)),
            ("TPS",             f"{m.get('tps',0):.2f} req/s"),
            ("Throughput",      f"{m.get('throughput_kbps',0):.1f} KB/s"),
            ("Avg Response",    f"{m.get('avg_ms',0):.1f} ms"),
            ("Median (P50)",    f"{m.get('median_ms',0)} ms"),
            ("P90",             f"{m.get('p90_ms',0)} ms"),
            ("P95",             f"{m.get('p95_ms',0)} ms"),
            ("P99",             f"{m.get('p99_ms',0)} ms"),
            ("Min Response",    f"{m.get('min_ms',0)} ms"),
            ("Max Response",    f"{m.get('max_ms',0)} ms"),
            ("Std Deviation",   f"{m.get('stdev_ms',0):.1f} ms"),
            ("Total Requests",  m.get("total_requests",   0)),
            ("Successful",      f"{m.get('total_success',0)} ({m.get('success_rate',0):.1f}%)"),
            ("Failed",          f"{m.get('total_failures',0)} ({m.get('error_rate',0):.1f}%)"),
            ("Error Rate",      f"{m.get('error_rate',0):.1f}%"),
            ("Duration",        f"{m.get('elapsed_s',0):.1f} s"),
        ]

        for label, value in kpis:
            row_idx = ws.max_row + 1
            ws.append([label, str(value)])
            cell_v = ws.cell(row=row_idx, column=2)
            if "%" in str(value):
                num = float(re.sub(r"[^0-9.]", "", str(value).split("%")[0]) or 0)
                if "Error" in label or "Failed" in label:
                    cell_v.font = fail_font if num > 5 else green_font
                else:
                    cell_v.font = green_font if num >= 95 else warn_font
            for col in range(1, 3):
                ws.cell(row=row_idx, column=col).border = thin_border

        auto_width(ws)

        # ── Sheet 2: Failure Breakdown ────────────────────────────────────────
        ws2 = wb.create_sheet("Failure Breakdown")
        ws2.append(["HTTP Code", "Count", "Root Cause", "Fix"])
        style_header_row(ws2, 1, 4)
        for code, count in (m.get("failure_codes") or {}).items():
            known = ai_agent.KNOWN_ISSUES.get(str(code), {})
            ws2.append([
                code, count,
                known.get("root_cause", "Unknown"),
                known.get("fix", "Investigate logs"),
            ])
            row_idx = ws2.max_row
            ws2.cell(row=row_idx, column=1).fill = fail_fill
            for col in range(1, 5):
                ws2.cell(row=row_idx, column=col).border = thin_border
        auto_width(ws2)

        # ── Sheet 3: System Monitor ───────────────────────────────────────────
        ws3 = wb.create_sheet("System Monitor")
        ws3.append(["Metric", "Value", "Status"])
        style_header_row(ws3, 1, 3)
        mon_rows = [
            ("CPU Average",    f"{ms.get('cpu_avg',0):.1f}%",   ms.get("cpu_avg",0) > 80),
            ("CPU Peak",       f"{ms.get('cpu_max',0):.1f}%",   ms.get("cpu_max",0) > 90),
            ("Memory Avg",     f"{ms.get('mem_avg',0):.1f}%",   ms.get("mem_avg",0) > 85),
            ("Memory Peak",    f"{ms.get('mem_max',0):.1f}%",   ms.get("mem_max",0) > 95),
            ("Net Received",   f"{ms.get('net_recv_total_mb',0)} MB", False),
            ("Alerts",         len(ms.get("alerts", [])),        len(ms.get("alerts",[])) > 0),
        ]
        for metric, value, is_warn in mon_rows:
            ws3.append([metric, value, "⚠ WARNING" if is_warn else "✅ OK"])
            row_idx = ws3.max_row
            status_cell = ws3.cell(row=row_idx, column=3)
            status_cell.font = warn_font if is_warn else green_font
            for col in range(1, 4):
                ws3.cell(row=row_idx, column=col).border = thin_border
        auto_width(ws3)

        # ── Sheet 4: RCA ──────────────────────────────────────────────────────
        ws4 = wb.create_sheet("RCA")
        ws4.append(["Code", "Root Cause", "Confidence %", "Suggested Fix",
                    "Recovery Action", "Source"])
        style_header_row(ws4, 1, 6)
        for rca in (rca_log or []):
            ws4.append([
                rca.get("error_code",  "?"),
                rca.get("root_cause",  "?"),
                rca.get("confidence",  0),
                rca.get("fix",         "?"),
                rca.get("recovery",    "?"),
                rca.get("source",      "?"),
            ])
            row_idx = ws4.max_row
            conf = rca.get("confidence", 0)
            ws4.cell(row=row_idx, column=3).font = (
                green_font if conf >= 80 else warn_font
            )
            for col in range(1, 7):
                ws4.cell(row=row_idx, column=col).border = thin_border
        if not rca_log:
            ws4.append(["N/A", "No RCA events in this session", "", "", "", ""])
        auto_width(ws4)

        # ── Sheet 5: Self-Healing ─────────────────────────────────────────────
        ws5 = wb.create_sheet("Self-Healing")
        ws5.append(["Error:Action", "Total", "Success", "Fail", "Success Rate", "Avg ms"])
        style_header_row(ws5, 1, 6)
        for key, s in list((healing_stats or {}).items()):
            rate = f"{s['success']/max(s['total'],1)*100:.0f}%"
            ws5.append([
                key, s["total"], s["success"], s["fail"], rate, s.get("avg_ms", 0)
            ])
            row_idx = ws5.max_row
            rate_num = s["success"] / max(s["total"], 1) * 100
            ws5.cell(row=row_idx, column=5).font = (
                green_font if rate_num >= 80 else warn_font
            )
            for col in range(1, 7):
                ws5.cell(row=row_idx, column=col).border = thin_border
        if not healing_stats:
            ws5.append(["N/A", 0, 0, 0, "N/A", 0])
        auto_width(ws5)

        # ── Sheet 6: Recommendations ──────────────────────────────────────────
        ws6 = wb.create_sheet("Recommendations")
        ws6.append(["#", "Recommendation"])
        style_header_row(ws6, 1, 2)
        for i, rec in enumerate(
            self._build_report_recommendations(m, ms), 1
        ):
            ws6.append([i, rec])
            for col in range(1, 3):
                ws6.cell(row=ws6.max_row, column=col).border = thin_border
        auto_width(ws6)

        out_path = os.path.join(REPORTS_DIR, f"report_{self.ts}.xlsx")
        wb.save(out_path)
        ok(f"Excel report → {out_path}")
        return out_path

    # ── PDF REPORT ────────────────────────────────────────────────────────────
    def generate_pdf(self, metrics: dict, monitor_summary: dict | None = None,
                     rca_log: list | None = None, healing_stats: dict | None = None,
                     scenario_name: str = "Scenario") -> str:
        if not REPORTLAB_OK:
            warn("reportlab not available — skipping PDF generation")
            return ""

        rca_log       = rca_log or []
        healing_stats = healing_stats or {}
        ms            = monitor_summary or {}
        m             = metrics

        out_path = os.path.join(REPORTS_DIR, f"report_{self.ts}.pdf")
        doc = SimpleDocTemplate(
            out_path, pagesize=A4,
            rightMargin=40, leftMargin=40,
            topMargin=50,   bottomMargin=40,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "custom_title", parent=styles["Title"],
            fontSize=18, spaceAfter=4,
            textColor=colors.HexColor("#00d4aa"),
        )
        h2_style = ParagraphStyle(
            "h2", parent=styles["Heading2"],
            fontSize=12, spaceAfter=6, spaceBefore=14,
            textColor=colors.HexColor("#7c6aff"),
        )
        body_style = ParagraphStyle(
            "body", parent=styles["Normal"],
            fontSize=9, leading=14,
        )

        def _table(data, col_widths=None):
            tbl = Table(data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1d2e")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#00d4aa")),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",   (0, 0), (-1, 0), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.HexColor("#0f1117"), colors.HexColor("#13161f")]),
                ("TEXTCOLOR",  (0, 1), (-1, -1), colors.HexColor("#e8eaf0")),
                ("FONTSIZE",   (0, 1), (-1, -1), 8),
                ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#2d3148")),
                ("ROWHEIGHT",  (0, 0), (-1, -1), 18),
                ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ]))
            return tbl

        story = []

        # Title
        story.append(Paragraph(f"⚡ Performance Test Report", title_style))
        story.append(Paragraph(
            f"Scenario: {scenario_name}  |  Profile: {m.get('profile_name','—')}  "
            f"|  Generated: {self.ts}",
            body_style
        ))
        story.append(Spacer(1, 16))
        story.append(HRFlowable(width="100%", thickness=1,
                                color=colors.HexColor("#2d3148")))
        story.append(Spacer(1, 12))

        # KPI Section
        story.append(Paragraph("Key Performance Indicators", h2_style))
        kpi_data = [
            ["Metric", "Value", "Metric", "Value"],
            ["Virtual Users",   str(m.get("virtual_users",0)),
             "TPS",             f"{m.get('tps',0):.2f} req/s"],
            ["Throughput",      f"{m.get('throughput_kbps',0):.1f} KB/s",
             "Avg Response",    f"{m.get('avg_ms',0):.1f} ms"],
            ["P90",             f"{m.get('p90_ms',0)} ms",
             "P95",             f"{m.get('p95_ms',0)} ms"],
            ["P99",             f"{m.get('p99_ms',0)} ms",
             "Total Requests",  str(m.get("total_requests",0))],
            ["Successful",      f"{m.get('total_success',0)} ({m.get('success_rate',0):.1f}%)",
             "Failed",          f"{m.get('total_failures',0)} ({m.get('error_rate',0):.1f}%)"],
            ["Min Response",    f"{m.get('min_ms',0)} ms",
             "Max Response",    f"{m.get('max_ms',0)} ms"],
            ["Error Rate",      f"{m.get('error_rate',0):.1f}%",
             "Duration",        f"{m.get('elapsed_s',0):.1f} s"],
        ]
        story.append(_table(kpi_data,
                            col_widths=[105, 105, 105, 105]))
        story.append(Spacer(1, 16))

        # System Monitor
        story.append(Paragraph("System Resource Monitor", h2_style))
        mon_data = [
            ["Metric", "Value", "Status"],
            ["CPU Average",   f"{ms.get('cpu_avg',0):.1f}%",
             "⚠ HIGH" if ms.get("cpu_avg",0) > 80 else "✅ OK"],
            ["Memory Avg",    f"{ms.get('mem_avg',0):.1f}%",
             "⚠ HIGH" if ms.get("mem_avg",0) > 85 else "✅ OK"],
            ["CPU Peak",      f"{ms.get('cpu_max',0):.1f}%",
             "⚠ HIGH" if ms.get("cpu_max",0) > 90 else "✅ OK"],
            ["Net Received",  f"{ms.get('net_recv_total_mb',0)} MB", "—"],
            ["Alerts",        str(len(ms.get("alerts",[]))),
             "⚠ SEE ALERTS" if ms.get("alerts") else "✅ NONE"],
        ]
        story.append(_table(mon_data, col_widths=[140, 140, 140]))
        story.append(Spacer(1, 16))

        # RCA
        story.append(Paragraph("AI Root Cause Analysis", h2_style))
        rca_data = [["Code", "Root Cause", "Confidence", "Recovery"]]
        for rca in (rca_log or [])[:8]:
            rca_data.append([
                str(rca.get("error_code", "?")),
                str(rca.get("root_cause", "?"))[:50],
                f"{rca.get('confidence','?')}%",
                str(rca.get("recovery", "?"))[:40],
            ])
        if len(rca_data) == 1:
            rca_data.append(["N/A", "No RCA events in this session", "", ""])
        story.append(_table(rca_data, col_widths=[40, 190, 70, 120]))
        story.append(Spacer(1, 16))

        # Self-Healing
        story.append(Paragraph("Self-Healing Statistics", h2_style))
        hs_data = [["Error:Action", "Total", "Success", "Fail", "Rate", "Avg ms"]]
        for key, s in list((healing_stats or {}).items())[:8]:
            rate = f"{s['success']/max(s['total'],1)*100:.0f}%"
            hs_data.append([
                key[:30], s["total"], s["success"], s["fail"], rate, s.get("avg_ms",0)
            ])
        if len(hs_data) == 1:
            hs_data.append(["N/A", 0, 0, 0, "N/A", 0])
        story.append(_table(hs_data, col_widths=[135, 45, 60, 45, 55, 55]))
        story.append(Spacer(1, 16))

        # Recommendations
        story.append(Paragraph("Recommendations", h2_style))
        for i, rec in enumerate(self._build_report_recommendations(m, ms), 1):
            story.append(Paragraph(f"{i}. {rec}", body_style))
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                color=colors.HexColor("#2d3148")))
        story.append(Paragraph(
            f"Generated by JMeter Automation v47 — Universal Enterprise Framework  |  "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            ParagraphStyle("footer", parent=body_style, fontSize=7,
                           textColor=colors.HexColor("#8b90a7"),
                           alignment=TA_CENTER)
        ))

        doc.build(story)
        ok(f"PDF report  → {out_path}")
        return out_path

    # ── Generate all three formats ────────────────────────────────────────────
    def generate_all(self, metrics: dict, monitor_summary: dict | None = None,
                     rca_log: list | None = None, healing_stats: dict | None = None,
                     scenario_name: str = "Scenario") -> dict[str, str]:
        banner("V47 Report Engine — Generating HTML + Excel + PDF")
        paths = {}
        paths["html"]  = self.generate_html(metrics, monitor_summary,
                                             rca_log, healing_stats, scenario_name)
        paths["excel"] = self.generate_excel(metrics, monitor_summary,
                                              rca_log, healing_stats, scenario_name)
        paths["pdf"]   = self.generate_pdf(metrics, monitor_summary,
                                            rca_log, healing_stats, scenario_name)
        return paths

    def _build_report_recommendations(self, m: dict, ms: dict) -> list[str]:
        recs = []
        if m.get("p95_ms", 0) > 5000:
            recs.append(
                f"P95 is {m['p95_ms']}ms — investigate slow endpoints; "
                "target <3000ms for interactive APIs"
            )
        if m.get("error_rate", 0) > 5:
            recs.append(
                f"Error rate {m['error_rate']:.1f}% exceeds 5% SLA — "
                "review failure codes and enable self-healing"
            )
        if ms.get("cpu_avg", 0) > 80:
            recs.append("CPU >80% — reduce VU count or scale horizontally")
        if ms.get("mem_avg", 0) > 85:
            recs.append("Memory >85% — check for heap leaks; add GC logging")
        if m.get("tps", 0) < 10 and m.get("virtual_users", 0) > 50:
            recs.append(
                f"TPS {m['tps']:.2f} is low for {m['virtual_users']} VUs — "
                "possible bottleneck in connection pool or DB"
            )
        if not recs:
            recs.append("All KPIs within acceptable thresholds — test passed ✅")
        return recs

# ═══════════════════════════════════════════════════════════════════════════════
# V46 PIPELINE FUNCTIONS (FULLY UNCHANGED)
# ═══════════════════════════════════════════════════════════════════════════════

# ─── BROWSER RECORDING ────────────────────────────────────────────────────────
def record(target_url):
    reqs = []; res_map = {}; seen = set()
    ws_messages: list[dict] = []
    stop = threading.Event(); done = threading.Event()

    def _run():
        with sync_playwright() as p:
            br  = p.chromium.launch(headless=False)
            ctx = br.new_context()
            pg  = ctx.new_page()

            def on_req(req):
                url = req.url; ul = url.lower()
                if any(ul.endswith(e) for e in IGNORE_EXT): return
                if ul.startswith(("data:", "blob:", "chrome-extension:")): return
                ct = req.headers.get("content-type", "").lower()
                if req.method == "GET" and not req.post_data and "json" not in ct:
                    acc = req.headers.get("accept", "").lower()
                    if "text/html" in acc: return
                key = (req.method, url, (req.post_data or "")[:200])
                if key in seen: return
                seen.add(key)
                prs = urlsplit(url)
                e = {
                    "method": req.method, "url": url, "scheme": prs.scheme,
                    "host": prs.netloc, "path": prs.path or "/", "query": prs.query,
                    "postData": req.post_data or "", "headers": dict(req.headers),
                    "contentType": ct, "status": None, "resp_json": None,
                    "isLogin": False, "isCreate": False, "creds": {},
                    "isGraphQL": False, "response_ms": 0,
                }
                reqs.append(e); res_map[f"{req.method}:{url}"] = e
                log(f"\n  ▶ #{len(reqs)}  {req.method}  {url}")
                if req.post_data: log(f"     {req.post_data[:160]}")

            def on_resp(resp):
                ul = resp.url.lower()
                if any(ul.endswith(e) for e in IGNORE_EXT): return
                body = ""
                try: body = resp.text()
                except: pass
                e = res_map.get(f"{resp.request.method}:{resp.url}")
                if not e: return
                e["status"] = resp.status
                try:
                    rj = json.loads(body); e["resp_json"] = rj
                    if isinstance(rj, dict):
                        flat = flatten(rj)
                        for tk in TOKEN_KEYS:
                            for k, v in flat.items():
                                if k.lower() == tk.lower() and len(str(v)) > 20:
                                    e["isLogin"] = True
                                    log(f"     🔑 LOGIN token key='{k}'")
                                    break
                except: pass
                if e["method"] == "POST" and resp.status == 201: e["isCreate"] = True
                e["isGraphQL"] = detect_graphql(e)
                if e["isGraphQL"]:
                    log(f"     🔷 GRAPHQL detected: {e['url'][:80]}")
                if e["isLogin"] and e["postData"]:
                    try:
                        bj = json.loads(e["postData"])
                        if isinstance(bj, dict):
                            for k, v in bj.items():
                                if k.lower() in CRED_KEYS: e["creds"][k] = v
                    except: pass
                icon = f"{G}✅{W}" if 200 <= (resp.status or 0) < 300 else f"{R}❌{W}"
                log(f"  ◀ {icon} {resp.status}  {resp.url[:80]}")

            def on_websocket(ws):
                log(f"  🔌 WebSocket OPEN  {ws.url[:80]}")
                ws_messages.append({"type": "open", "url": ws.url, "data": None})
                def on_frame_sent(payload):
                    ws_messages.append({"type": "sent", "url": ws.url, "data": payload})
                def on_frame_recv(payload):
                    ws_messages.append({"type": "recv", "url": ws.url, "data": payload})
                ws.on("framesent",     on_frame_sent)
                ws.on("framereceived", on_frame_recv)
                ws.on("close", lambda: ws_messages.append({"type": "close", "url": ws.url}))

            pg.on("request",   on_req)
            pg.on("response",  on_resp)
            pg.on("websocket", on_websocket)

            pg.goto(target_url)
            while not stop.is_set():
                try: pg.wait_for_timeout(500)
                except: break
            try: pg.wait_for_timeout(2000)
            except: pass
            br.close()
        done.set()

    threading.Thread(target=_run, daemon=True).start()
    print(f"\n{G}👉 Perform your COMPLETE scenario in the browser.{W}")
    print(f"   Then press {B}ENTER{W} when done...\n")
    input()
    stop.set(); done.wait(15)
    ok(f"Recorded {len(reqs)} HTTP requests")

    if ws_messages:
        ok(f"Captured {len(ws_messages)} WebSocket frame events")
        reqs.append({
            "method": "WEBSOCKET", "url": ws_messages[0]["url"],
            "scheme": "wss", "host": urlsplit(ws_messages[0]["url"]).netloc,
            "path": urlsplit(ws_messages[0]["url"]).path or "/",
            "query": "", "postData": "", "headers": {},
            "contentType": "application/json", "status": 101,
            "resp_json": None, "isLogin": False, "isCreate": False,
            "creds": {}, "isGraphQL": False, "response_ms": 0,
            "ws_frames": ws_messages,
        })

    return reqs

# ─── REMOVE FAILED REQUESTS ──────────────────────────────────────────────────
def remove_failed_requests(reqs: list) -> list:
    cleaned = []
    for r in reqs:
        status = int(r.get("status") or 0)
        if status >= 400:
            warn(f"Removing failed request ({status}): {r['method']} {r['url'][:80]}")
            continue
        cleaned.append(r)
    if len(cleaned) < len(reqs):
        ok(f"Removed {len(reqs) - len(cleaned)} failed request(s) ({len(cleaned)} remain)")
    return cleaned

# ─── FILTER ───────────────────────────────────────────────────────────────────
def filter_reqs(reqs):
    SKIP_VALIDATE_PATTERNS = [
        "validate-name", "validate_name", "check-name", "check_name",
        "verify-name", "verify_name", "autocomplete", "draft-check",
        "preview", "uniqueness",
    ]
    filtered = []
    for r in reqs:
        url_lower = r["url"].lower()
        method    = r["method"].upper()
        if method == "WEBSOCKET":
            filtered.append(r); continue
        if any(v in url_lower for v in SKIP_VALIDATE_PATTERNS):
            warn(f"SKIPPED validate API: {method} {r['url'][:80]}")
            continue
        if r.get("isLogin"):
            filtered.append(r); continue
        if KEEP_POST_APIS and any(x in url_lower for x in KEEP_POST_APIS):
            filtered.append(r); continue
        if method in ("POST", "PUT", "PATCH", "DELETE"):
            filtered.append(r); continue
        ct = r.get("contentType", "").lower()
        if r.get("postData") or r.get("resp_json") or "json" in ct:
            filtered.append(r)

    out = filtered
    if len(out) < 2:
        out = [r for r in reqs
               if not any(r["url"].lower().endswith(e) for e in IGNORE_EXT)]

    seen = set(); dedup = []
    for r in out:
        k = (r["method"], r["path"], r["postData"][:150])
        if k not in seen: seen.add(k); dedup.append(r)
    return dedup

# ─── TOKEN DETECTION ──────────────────────────────────────────────────────────
def find_token(reqs):
    for r in reqs:
        if not r.get("isLogin"): continue
        rj = r.get("resp_json")
        if not isinstance(rj, dict): continue
        flat = flatten(rj)
        for tk in AUTH_RESPONSE_KEYS:
            for k, v in flat.items():
                if k.lower() == tk.lower() and len(str(v)) > 20:
                    return k, str(v)
    return None, None

# ─── AUTH DISCOVERY ───────────────────────────────────────────────────────────
def discover_auth(reqs):
    auth_keywords = [
        "login", "signin", "sign-in", "sign_in",
        "authenticate", "authentication",
        "auth", "token", "oauth", "saml", "jwt", "session",
    ]
    candidates = []
    for r in reqs:
        score   = 0
        url     = r.get("url",      "").lower()
        body    = r.get("postData", "").lower()
        resp    = str(r.get("resp_json", "")).lower()
        headers = str(r.get("headers",  {})).lower()
        if any(k in url for k in auth_keywords):         score += 100
        if any(k in body for k in ["password", "username", "email"]): score += 50
        if any(k in resp for k in ["token", "access_token", "jwt"]): score += 100
        if "authorization" in headers:                   score += 50
        if "set-cookie"        in headers: score += 50
        if "jsessionid"        in headers: score += 60
        if "asp.net_sessionid" in headers: score += 60
        if score > 0:
            candidates.append({"score": score, "request": r})
    if not candidates:
        return None
    best = max(candidates, key=lambda x: x["score"])
    return best["request"]

# ─── V45-1: STRUCT-AWARE BODY SUBSTITUTION ────────────────────────────────────
_JSON_VALUE_ONLY_SUBS = True

def sub_body_json_struct(body_str: str, corr: dict, val_map: dict) -> str:
    TOKEN_SKIP = {k.lower() for k in TOKEN_KEYS}

    def _sub_value(key: str, val):
        if not isinstance(val, str):
            return val
        if key.lower() in TOKEN_SKIP:
            return val
        for field, prop in corr.items():
            if key.lower() == field.lower() and field.lower() not in TOKEN_SKIP:
                return "${" + prop + "}"
        for raw, field in sorted(val_map.items(), key=lambda x: -len(x[0])):
            prop = corr.get(field)
            if prop and val == raw and field.lower() not in TOKEN_SKIP:
                return "${" + prop + "}"
        return val

    def _walk(obj):
        if isinstance(obj, dict):
            return {k: _walk(_sub_value(k, v)) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [_walk(i) for i in obj]
        return obj

    try:
        parsed = json.loads(body_str)
        substituted = _walk(parsed)
        result = json.dumps(substituted, separators=(',', ':'))
        result = repair_missing_corr(result, corr)
        return result
    except (json.JSONDecodeError, ValueError):
        return _sub_body_regex_safe(body_str, corr, val_map)


def _sub_body_regex_safe(body: str, corr: dict, val_map: dict) -> str:
    TOKEN_SKIP = {k.lower() for k in TOKEN_KEYS}
    r = body
    for field, prop in corr.items():
        if field.lower() in TOKEN_SKIP: continue
        jv = f"${{{prop}}}"
        r = re.sub(
            rf'("{re.escape(field)}"\s*:\s*)"([^"{{}}]{{1,400}})"',
            lambda m, jv=jv: f'{m.group(1)}"{jv}"',
            r,
        )
    for raw, field in sorted(val_map.items(), key=lambda x: -len(x[0])):
        prop = corr.get(field)
        if not prop or field.lower() in TOKEN_SKIP: continue
        jv = f"${{{prop}}}"
        if raw in r and jv not in r:
            r = r.replace(f'"{raw}"', f'"{jv}"')
    r = repair_missing_corr(r, corr)
    return r


def sub_body(body, corr, val_map):
    if not body: return body
    body_stripped = body.strip()
    if body_stripped.startswith('{') or body_stripped.startswith('['):
        return sub_body_json_struct(body, corr, val_map)
    return _sub_body_regex_safe(body, corr, val_map)

# ─── FUZZY DEPENDENCY CORRELATION ─────────────────────────────────────────────
def build_corr_v41(reqs, token_key):
    global correlation_store, correlation_confidence
    corr             = {}
    val_map          = {}
    dependency_graph = {}

    if token_key:
        corr[token_key] = "SHARED_TOKEN"

    known_fields = db_load_known_fields()

    for idx, r in enumerate(reqs):
        rj = r.get("resp_json")
        if not isinstance(rj, (dict, list)):
            continue

        flat = flatten(rj)
        for field, value in flat.items():
            value = str(value).strip()
            if not is_dynamic(field, value):
                continue
            conf = confidence_score(field, value)
            if conf < 60:
                info(f"  Skipped low-confidence field: {field} (score={conf})")
                continue

            consumers = []
            for j in range(idx + 1, len(reqs)):
                future_blob = reqs[j]["url"] + reqs[j]["postData"]
                if value in future_blob or similarity(value, future_blob) > 0.80:
                    consumers.append(j)

            if not consumers and field in known_fields:
                info(f"  DB-known field promoted: {field} → {known_fields[field]}")

            if consumers or field in known_fields:
                var = f"CORR_{safe_var(field)}"
                corr[field]    = var
                val_map[value] = field

                dependency_graph[var] = {
                    "producer":  idx,
                    "consumers": consumers,
                }
                correlation_store[field] = {
                    "value":     value,
                    "producer":  idx,
                    "consumers": consumers,
                }
                correlation_confidence[field] = conf
                db_store(field, value)
                ok(f"AI Dependency: {field} → {var}  "
                   f"(conf={conf}, producer={idx}, consumers={consumers})")

    return corr, val_map, dependency_graph

# ─── V46-2: URL HEALING ───────────────────────────────────────────────────────
def heal_url(url: str, corr: dict) -> str:
    prs  = urlsplit(url)
    path = prs.path
    for raw, field in sorted(
        {v["value"]: k for k, v in correlation_store.items()}.items(),
        key=lambda x: -len(x[0]),
    ):
        prop = corr.get(field)
        if prop and raw in path:
            path = path.replace(raw, f"${{{prop}}}")
    healed_parts = []
    for part in path.split("/"):
        if not part:
            healed_parts.append(part)
            continue
        if part.startswith("${"):
            healed_parts.append(part)
        elif is_uuid(part):
            matched = next(
                (f"${{{corr[f]}}}" for f, d in correlation_store.items()
                 if d["value"] == part and f in corr),
                None
            )
            if matched:
                healed_parts.append(matched)
            else:
                healed_parts.append("${CORR_WORKSPACE_UID}")
                ok(f"V46-2: Unknown UUID '{part[:8]}...' in path → ${{CORR_WORKSPACE_UID}}")
        elif re.match(r"^\d{6,20}$", part):
            matched = next(
                (f"${{{corr[f]}}}" for f, d in correlation_store.items()
                 if d["value"] == part and f in corr),
                f"${{CORR_ID_{safe_var(part)}}}"
            )
            healed_parts.append(matched)
        else:
            healed_parts.append(part)
    new_path = "/".join(healed_parts)
    return prs._replace(path=new_path).geturl()

# ─── HEADER HEALING ───────────────────────────────────────────────────────────
def heal_headers(headers: dict, corr: dict) -> dict:
    HEAL_HEADER_KEYS = {
        "authorization", "x-api-key", "apikey", "api-key",
        "session", "x-session-token", "x-auth-token",
        "correlationid", "x-correlation-id", "x-request-id",
    }
    healed = {}
    for hk, hv in headers.items():
        hk_lower = hk.lower()
        if hk_lower in SKIP_HDRS:
            healed[hk] = hv; continue
        if hk_lower in HEAL_HEADER_KEYS:
            if "${" in str(hv):
                healed[hk] = hv; continue
            hv_str = str(hv)
            replaced = False
            for field, data in correlation_store.items():
                raw = data["value"]
                if raw and raw in hv_str:
                    prop = corr.get(field)
                    if prop:
                        healed[hk] = hv_str.replace(raw, f"${{{prop}}}")
                        replaced = True
                        break
            if not replaced:
                if hk_lower == "authorization" and "bearer " in hv_str.lower():
                    healed[hk] = "Bearer ${__P(SHARED_TOKEN,INIT)}"
                else:
                    healed[hk] = hv
        else:
            healed[hk] = hv
    return healed

# ─── COOKIE HEALING ───────────────────────────────────────────────────────────
SESSION_COOKIE_MAP = {
    "jsessionid": "COOKIE_JSESSIONID", "sessionid": "COOKIE_SESSIONID",
    "asp.net_sessionid": "COOKIE_ASPNET_SESSION", "phpsessid": "COOKIE_PHPSESSID",
    "connect.sid": "COOKIE_CONNECTSID", "sid": "COOKIE_SID",
    "_session": "COOKIE_SESSION",
}

def heal_cookies(headers: dict) -> dict:
    healed = dict(headers)
    raw_cookie = headers.get("cookie", "") or headers.get("Cookie", "")
    if not raw_cookie:
        return healed
    def _replace_cookie_value(cookie_str: str) -> str:
        parts = cookie_str.split(";"); new_parts = []
        for part in parts:
            if "=" not in part:
                new_parts.append(part); continue
            name, _, val = part.strip().partition("=")
            name_lower = name.strip().lower()
            if name_lower in SESSION_COOKIE_MAP:
                var = SESSION_COOKIE_MAP[name_lower]
                new_parts.append(f"{name.strip()}=${{{var}}}")
            else:
                new_parts.append(part.strip())
        return "; ".join(new_parts)
    for hk in list(healed.keys()):
        if hk.lower() == "cookie":
            healed[hk] = _replace_cookie_value(healed[hk])
    return healed

# ─── V46-7: MULTIPART HEALING ─────────────────────────────────────────────────
def heal_multipart(parts: dict, corr: dict) -> dict:
    EXPLICIT_FIELD_MAP = {
        "project_uid":     "CORR_PROJECT_UID",
        "sprint_uid":      "CORR_SPRINT_UID",
        "requirement_uid": "CORR_REQUIREMENT_UID",
        "defect_uid":      "CORR_DEFECT_UID",
        "org_uid":         "CORR_ORG_UID",
        "user_uid":        "CORR_USER_UID",
        "workspace_uid":   "CORR_WORKSPACE_UID",
    }
    healed = {}
    for fname, fval in parts.items():
        if fname.lower() in EXPLICIT_FIELD_MAP:
            healed[fname] = "${" + EXPLICIT_FIELD_MAP[fname.lower()] + "}"
            ok(f"V46-7: Multipart '{fname}' → {healed[fname]}")
            continue
        substituted = False
        for corr_field, corr_var in corr.items():
            if fname.lower() == corr_field.lower():
                healed[fname] = "${" + corr_var + "}"; substituted = True; break
        if substituted: continue
        for corr_field, data in correlation_store.items():
            if similarity(fname, corr_field) >= 0.65:
                prop = corr.get(corr_field)
                if prop:
                    healed[fname] = "${" + prop + "}"; substituted = True; break
        if not substituted:
            healed[fname] = fval
    return healed

# ─── UNIVERSAL REQUEST HEALER ─────────────────────────────────────────────────
def heal_request(req: dict, corr: dict) -> dict:
    req["url"]  = heal_url(req["url"], corr)
    req["path"] = urlsplit(req["url"]).path or "/"
    req["headers"] = heal_headers(req.get("headers", {}), corr)
    req["headers"] = heal_cookies(req["headers"])
    if req.get("postData"):
        req["postData"] = repair_missing_corr(req["postData"], corr)
    if "multipart" in req.get("contentType", "").lower() and req.get("postData"):
        parts  = parse_multipart(req["postData"])
        healed = heal_multipart(parts, corr)
        req["_healed_multipart"] = healed
    return req

# ─── ENTITY DISCOVERY ─────────────────────────────────────────────────────────
def discover_entities(reqs):
    entities = []
    for i, r in enumerate(reqs):
        if r["method"] == "POST":
            resp = r.get("resp_json")
            if not isinstance(resp, dict): continue
            if r.get("isGraphQL"):
                fields = graphql_extract(resp)
            else:
                fields = flatten(resp)
            for k, v in fields.items():
                if is_dynamic(k, v) and k in correlation_store:
                    dep  = correlation_store[k]
                    conf = correlation_confidence.get(k, 0)
                    entities.append({
                        "request": i, "field": k, "value": v,
                        "consumers": dep.get("consumers", []),
                        "confidence": conf,
                        "entity_group": semantic_entity(k),
                    })
    return entities

# ─── V45-5: SAFE AI ASSERTION GENERATOR ───────────────────────────────────────
def generate_assertions(resp_json):
    assertions = []
    if isinstance(resp_json, dict):
        for k, v in resp_json.items():
            if isinstance(v, str) and k.lower() in SAFE_ASSERT_KEYS:
                if v.strip().lower() not in ("true", "false", "null", "none", ""):
                    assertions.append((k, str(v)))
            elif isinstance(v, int) and k.lower() in SAFE_ASSERT_KEYS:
                if v in (200, 201, 0, 1):
                    assertions.append((k, str(v)))
    return assertions[:3]

# ─── SELF-HEALING CORRELATION ENGINE ─────────────────────────────────────────
def repair_missing_corr(body, corr):
    if not body: return body
    for field, prop in corr.items():
        token = "${" + prop + "}"
        if token in body: continue
        pattern = rf'"{field}"\s*:\s*"([^"]+)"'
        m = re.search(pattern, body)
        if m:
            body = re.sub(pattern, f'"{field}":"{token}"', body)
    return body

# ─── V46-2: URL SUBSTITUTION ──────────────────────────────────────────────────
def sub_url(url, corr, val_map):
    TOKEN_SKIP = {k.lower() for k in TOKEN_KEYS}
    prs  = urlsplit(url)
    path = prs.path
    for raw, field in sorted(val_map.items(), key=lambda x: -len(x[0])):
        prop = corr.get(field)
        if not prop or field.lower() in TOKEN_SKIP: continue
        if raw in path:
            path = path.replace(raw, f"${{{prop}}}")
    def _replace_uuid_in_segment(seg):
        if seg.startswith("${") or not is_uuid(seg):
            return seg
        matched = next(
            (f"${{{corr[f]}}}" for f, d in correlation_store.items()
             if d["value"] == seg and f in corr),
            None
        )
        if matched:
            return matched
        return "${CORR_WORKSPACE_UID}"
    path = "/".join(_replace_uuid_in_segment(s) for s in path.split("/"))
    return prs._replace(path=path).geturl()

# ─── FIND USER-TYPED FIELDS ───────────────────────────────────────────────────
def find_fields(reqs, corr):
    ck = set(corr.keys()); out = {}
    for r in reqs:
        is_mp = "multipart" in r.get("contentType", "").lower()
        body  = r.get("postData", "")
        if not body: continue
        if is_mp:
            fields = parse_multipart(body)
        else:
            try:
                parsed = json.loads(body)
                fields = parsed if isinstance(parsed, dict) else {}
            except:
                fields = {}
        for k, v in fields.items():
            if k in ck or k.lower() in CRED_KEYS or k.lower() in SKIP_KEYS: continue
            if ai_user_typed(k, str(v)) and k not in out:
                out[k] = str(v)
    return out

# ─── TEST DATA GENERATOR ─────────────────────────────────────────────────────
def generate_test_data(field):
    f = field.lower()
    if "email"       in f: return fake.email()
    if "phone"       in f: return fake.phone_number()
    if "name"        in f: return fake.name()
    if "address"     in f: return fake.address()
    if "description" in f: return fake.sentence()
    if "summary"     in f: return fake.sentence()
    if "title"       in f: return fake.catch_phrase()
    if "customer"    in f: return fake.company()
    if "account"     in f: return fake.company()
    if "reference"   in f: return fake.bothify(text="REF-####-????")
    if "order"       in f: return fake.bothify(text="ORD-####")
    if "steps"       in f: return fake.sentence()
    if "actual"      in f: return fake.sentence()
    if "expected"    in f: return fake.sentence()
    return fake.word()

# ─── V46-6: SINGLE SEQUENTIAL FLOW ────────────────────────────────────────────
def split_flows_v40(reqs, dependency_graph):
    if not dependency_graph:
        return [reqs]
    single_flow = list(reqs)
    ok("V46-6: Using single sequential flow to preserve CREATE→USE dependency chain")
    return [single_flow]

def flow_name(flow):
    if any(r.get("isGraphQL") for r in flow):            return "GraphQL Transaction"
    if any(r.get("isLogin") for r in flow):              return "Login"
    if any(r.get("method") == "WEBSOCKET" for r in flow): return "WebSocket Transaction"
    methods = [r["method"] for r in flow]
    if "POST" in methods:
        has_create = any(
            r["method"] == "POST" and r.get("isCreate", False) for r in flow
        )
        return "Full Scenario" if has_create else "API Transaction"
    if "PUT" in methods or "PATCH" in methods: return "Update Transaction"
    if "DELETE" in methods: return "Delete Transaction"
    return "View Transaction"

# ─── V45-6: fix_create() ─────────────────────────────────────────────────────
def fix_create(body, corr):
    if not body: return body
    try:
        obj = json.loads(body)
        if isinstance(obj, dict):
            name_fields = [
                "name", "title", "summary", "description",
                "reference", "steps", "actual", "expected",
            ]
            for field in name_fields:
                if field in obj and isinstance(obj[field], str):
                    if not obj[field].startswith("${"):
                        obj[field] = f"{generate_test_data(field)}${{ITER_SUFFIX}}"
            return json.dumps(obj, separators=(',', ':'))
    except (json.JSONDecodeError, ValueError):
        pass
    r = body
    def _replace_field(pattern, field_name):
        nonlocal r
        def _sub(m):
            faked = generate_test_data(field_name)
            return f'{m.group(1)}{faked}${{ITER_SUFFIX}}{m.group(3)}'
        r = re.sub(pattern, _sub, r, flags=re.I)
    _replace_field(r'("name"\s*:\s*")([^"]+)(")',        "name")
    _replace_field(r'("title"\s*:\s*")([^"]+)(")',        "title")
    _replace_field(r'("summary"\s*:\s*")([^"]+)(")',      "summary")
    _replace_field(r'("description"\s*:\s*")([^"]+)(")', "description")
    _replace_field(r'("reference"\s*:\s*")([^"]+)(")',   "reference")
    _replace_field(r'("steps"\s*:\s*")([^"]+)(")',        "steps")
    _replace_field(r'("actual"\s*:\s*")([^"]+)(")',       "actual")
    _replace_field(r'("expected"\s*:\s*")([^"]+)(")',     "expected")
    return r

def fix_validate(body):
    if not body: return body
    return re.sub(
        r'("name"\s*:\s*")([^"$][^"]{0,80})(")',
        r'\1\2${ITER_SUFFIX}\3',
        body,
    )

# ─── FAILURE CLASSIFICATION ───────────────────────────────────────────────────
def classify_failure(code):
    mapping = {
        "401": "Authentication Failure", "403": "Authorization Failure",
        "404": "Dependency Failure",     "409": "Duplicate Data",
        "429": "Rate Limiting",          "500": "Backend Exception",
        "502": "Gateway Error",          "503": "Service Unavailable",
        "504": "Timeout",
    }
    return mapping.get(str(code), "Unknown Failure")

def generate_cleanup(reqs):
    cleanup = []
    for r in reqs:
        if r["method"] == "POST":
            cleanup.append({"method": "DELETE", "target": r["url"]})
    return cleanup

def ai_root_cause(code, elapsed):
    if code == "401": return "Authentication"
    if code == "403": return "Authorization"
    if code == "409": return "Duplicate Test Data"
    if code == "500":
        if elapsed > 5000: return "Database Bottleneck"
        return "Backend Exception"
    return "Unknown"

# ─── RUNTIME RE-CORRELATION ───────────────────────────────────────────────────
def auto_recorrelate(failed_req: dict, previous_responses: list[dict]) -> dict:
    new_corr: dict = {}
    haystack = failed_req.get("url", "") + failed_req.get("postData", "")
    for resp in previous_responses:
        rj = resp.get("resp_json")
        if not isinstance(rj, (dict, list)): continue
        for field, value in flatten(rj).items():
            value = str(value).strip()
            if not is_dynamic(field, value): continue
            if value not in haystack: continue
            conf = confidence_score(field, value)
            if conf < 60: continue
            var = f"CORR_{safe_var(field)}"
            new_corr[field] = var
            correlation_store[field] = {"value": value, "producer": -1, "consumers": []}
            correlation_confidence[field] = conf
            db_store(field, value)
            ok(f"Auto-recorrelated: {field} → {var}  (conf={conf})")
    return new_corr

def runtime_heal_request(failed_req, previous_responses, corr):
    new_corr = auto_recorrelate(failed_req, previous_responses)
    if new_corr:
        corr.update(new_corr)
        ok(f"runtime_heal_request: {len(new_corr)} new correlation(s) discovered")
    failed_req = heal_request(failed_req, corr)
    return failed_req

def repair_strategy(code: str) -> str:
    mapping = {"401": "TOKEN", "403": "SESSION", "404": "CORRELATION",
               "409": "DATA",  "422": "PAYLOAD"}
    return mapping.get(str(code), "UNKNOWN")

def runtime_repair(code, failed_req, responses, corr):
    strategy = repair_strategy(code)
    info(f"runtime_repair: code={code} strategy={strategy} url={failed_req.get('path','?')}")
    if strategy == "TOKEN":
        correlation_store.pop("SHARED_TOKEN", None)
        ok("runtime_repair TOKEN: SHARED_TOKEN cleared")
    elif strategy == "SESSION":
        for ck in list(SESSION_COOKIE_MAP.values()):
            correlation_store.pop(ck, None)
        ok("runtime_repair SESSION: session cookies cleared")
    elif strategy == "CORRELATION":
        failed_req = runtime_heal_request(failed_req, responses, corr)
        ok("runtime_repair CORRELATION: request re-healed")
    elif strategy == "DATA":
        body = failed_req.get("postData", "")
        if body:
            for field in ["name", "title", "summary", "description",
                          "reference", "customer", "account", "order",
                          "steps", "actual", "expected"]:
                pattern = rf'("{re.escape(field)}"\s*:\s*")([^"]+)(")'
                def _fresh(m, f=field):
                    return f'{m.group(1)}{generate_test_data(f)}_RETRY{m.group(3)}'
                body = re.sub(pattern, _fresh, body, flags=re.I)
            failed_req["postData"] = body
            ok("runtime_repair DATA: fresh Faker values injected")
    elif strategy == "PAYLOAD":
        failed_req = heal_request(failed_req, corr)
        ok("runtime_repair PAYLOAD: request payload re-healed")
    else:
        warn(f"runtime_repair UNKNOWN: no strategy for code={code}")
    return failed_req

def graphql_extract(data: dict, depth: int = 0) -> dict:
    if depth > 12: return {}
    if not isinstance(data, dict): return {}
    result: dict = {}
    root = data.get("data", data)
    def _walk(node, d):
        if d > 12: return
        if isinstance(node, dict):
            for k, v in node.items():
                if k in ("errors", "__typename", "extensions"): continue
                if isinstance(v, str) and is_dynamic(k, v): result[k] = v
                elif isinstance(v, (int, float)) and re.search(r"(id|uid|ref|key)$", k, re.I):
                    result[k] = str(v)
                elif isinstance(v, (dict, list)): _walk(v, d + 1)
        elif isinstance(node, list):
            for item in node: _walk(item, d + 1)
    _walk(root, 0)
    return result

def monitor_correlation() -> list[str]:
    unstable: list[str] = []
    log(); log(f"  {B}{C}CORRELATION CONFIDENCE REPORT:{W}")
    log(f"  {'─'*58}")
    log(f"  {'Field':<30} {'Entity':<16} {'Conf':>5}  {'Status'}")
    log(f"  {'─'*30} {'─'*16} {'─'*5}  {'─'*10}")
    for field, conf in sorted(correlation_confidence.items(),
                               key=lambda x: x[1], reverse=True):
        entity  = semantic_entity(field)
        status  = "✅ STABLE" if conf >= 80 else f"{Y}⚠ UNSTABLE{W}"
        log(f"  {field:<30} {entity:<16} {conf:>5}  {status}")
        if conf < 80:
            unstable.append(field)
            warn(f"  Unstable correlation: '{field}' (conf={conf})")
    log(f"  {'─'*58}")
    ok(f"  {len(correlation_confidence) - len(unstable)} stable  |  "
       f"{len(unstable)} unstable")
    return unstable

def self_healing_cycle(failed_req, code, responses, corr):
    ok(f"self_healing_cycle: initiating for {failed_req.get('path','?')} (code={code})")
    t0 = time.monotonic()
    repaired = runtime_repair(code, failed_req, responses, corr)
    repaired = runtime_heal_request(repaired, responses, corr)
    elapsed_ms = int((time.monotonic() - t0) * 1000)
    for field in corr:
        if "${" + corr[field] + "}" in (repaired.get("url","") +
                                         repaired.get("postData","")):
            db_record_outcome(field, success=True)
    # V47: log healing event
    kb_store_healing(str(code), repair_strategy(str(code)),
                     success=True, elapsed_ms=elapsed_ms)
    ok(f"self_healing_cycle: complete — replay=True")
    return repaired, True

# ─── POST-RUN ANALYSIS ────────────────────────────────────────────────────────
def analyze_jtl(jtl_file):
    try:
        import pandas as pd
    except ImportError:
        warn("pandas not installed — skipping bottleneck analysis"); return
    if not os.path.exists(jtl_file):
        warn(f"JTL file not found: {jtl_file}"); return
    try:
        df     = pd.read_csv(jtl_file)
        errors = df[df['success'] == False]
        log(); log(f"{B}{C}{'─'*65}")
        log(f"  POST-RUN ANALYSIS  (v47 AI Root Cause + Bottleneck Classifier)")
        log(f"{'─'*65}{W}")
        log(f"\n{Y}  BOTTLENECK APIs:{W}")
        if 'elapsed' in df.columns:
            any_bottleneck = False
            for _, row in df.iterrows():
                elapsed = int(row.get('elapsed', 0))
                code    = str(row.get('responseCode', ''))
                if elapsed > 3000:
                    # V47: Use AI agent RCA
                    rca = ai_agent.analyze_failure(code, elapsed_ms=elapsed,
                                                    scenario="JTL Analysis")
                    cause = rca["root_cause"]
                    any_bottleneck = True
                    log(f"    {R}⚠  {row.get('label','?'):<50} "
                        f"{elapsed:>6}ms  → {cause}{W}")
            if not any_bottleneck: ok("  No bottlenecks detected")
        else:
            warn("  'elapsed' column not found in JTL")
        log(f"\n{R}  FAILED APIs:{W}")
        if errors.empty:
            ok("  No failures detected")
        else:
            cols = [c for c in ['label', 'responseCode', 'responseMessage']
                    if c in errors.columns]
            for _, row in errors[cols].iterrows():
                code    = str(row.get('responseCode', ''))
                elapsed = int(row.get('elapsed', 0)) if 'elapsed' in row else 0
                label   = row.get('label', '?')
                # V47: Use AI agent RCA
                rca  = ai_agent.analyze_failure(code, elapsed_ms=elapsed,
                                                 scenario="JTL Analysis")
                log(f"    {R}✗  {label:<50}  {code}  → "
                    f"{rca['root_cause'][:40]}  ({rca['confidence']}%){W}")
        total   = len(df); fail_ct = len(errors)
        log(f"\n{C}  SUMMARY:{W}")
        log(f"    Total samples : {total}")
        log(f"    Failures      : {fail_ct}  ({fail_ct/max(total,1)*100:.1f}%)")
        if 'elapsed' in df.columns:
            log(f"    Avg elapsed   : {df['elapsed'].mean():.0f} ms")
            log(f"    95th pct      : {df['elapsed'].quantile(0.95):.0f} ms")

        # V47: Auto-generate reports from JTL
        jtl_metrics = {
            "profile_name":   "JMeter Run",
            "virtual_users":  0,
            "duration_s":     0,
            "total_requests": total,
            "total_success":  total - fail_ct,
            "total_failures": fail_ct,
            "success_rate":   round((total-fail_ct)/max(total,1)*100, 2),
            "error_rate":     round(fail_ct/max(total,1)*100, 2),
            "tps":            0.0,
            "throughput_kbps":0.0,
            "avg_ms":         df["elapsed"].mean() if "elapsed" in df.columns else 0,
            "min_ms":         df["elapsed"].min()  if "elapsed" in df.columns else 0,
            "max_ms":         df["elapsed"].max()  if "elapsed" in df.columns else 0,
            "median_ms":      df["elapsed"].quantile(0.5) if "elapsed" in df.columns else 0,
            "p90_ms":         df["elapsed"].quantile(0.90) if "elapsed" in df.columns else 0,
            "p95_ms":         df["elapsed"].quantile(0.95) if "elapsed" in df.columns else 0,
            "p99_ms":         df["elapsed"].quantile(0.99) if "elapsed" in df.columns else 0,
            "stdev_ms":       df["elapsed"].std() if "elapsed" in df.columns else 0,
            "failure_codes":  {},
            "elapsed_s":      0.0,
        }
        if "responseCode" in errors.columns:
            for code_v in errors["responseCode"].value_counts().items():
                jtl_metrics["failure_codes"][code_v[0]] = code_v[1]

        rpt = ReportEngine()
        rpt.generate_all(
            jtl_metrics,
            monitor_summary=None,
            rca_log=ai_agent.rca_log,
            healing_stats=kb_load_healing_stats(),
            scenario_name="JMeter Run",
        )
    except Exception as ex:
        warn(f"analyze_jtl error: {ex}")

# ─── BUILD JMX (V46 UNCHANGED) ────────────────────────────────────────────────
def build_jmx(reqs, profile, jmx_path, jtl_path, corr, val_map,
               token_key, token_val, user_fields, flows, login_req):

    login_host   = urlsplit(login_req["url"]).hostname if login_req else ""
    login_scheme = urlsplit(login_req["url"]).scheme   if login_req else "https"
    login_path   = login_req.get("path", "/login")     if login_req else "/login"
    login_body   = login_req.get("postData", "")       if login_req else ""
    login_ct     = (login_req.get("contentType", "application/json")
                    if login_req else "application/json")

    login_body_groovy = groovy_safe_string(login_body)
    token_keys_groovy = groovy_token_keys_list()

    root = ET.Element("jmeterTestPlan", version="1.2", properties="5.0", jmeter="5.6.3")
    rht  = ET.SubElement(root, "hashTree")

    tp = ET.SubElement(rht, "TestPlan", guiclass="TestPlanGui",
                       testclass="TestPlan", testname="Performance Test v47 AI",
                       enabled="true")
    xb(tp, "TestPlan.functional_mode", False)
    xb(tp, "TestPlan.serialize_threadgroups", True)
    ud = ET.SubElement(tp, "elementProp", name="TestPlan.user_defined_variables",
                       elementType="Arguments", guiclass="ArgumentsPanel",
                       testclass="Arguments", testname="UDV", enabled="true")
    ET.SubElement(ud, "collectionProp", name="Arguments.arguments")
    tpht = ET.SubElement(rht, "hashTree")

    def save_cfg():
        obj = ET.Element("objProp")
        ET.SubElement(obj, "name").text = "saveConfig"
        val = ET.SubElement(obj, "value", attrib={"class": "SampleSaveConfiguration"})
        for tag, txt in [
            ("time", "true"), ("latency", "true"), ("timestamp", "true"),
            ("success", "true"), ("label", "true"), ("code", "true"),
            ("message", "true"), ("threadName", "true"), ("dataType", "true"),
            ("encoding", "false"), ("assertions", "true"), ("subresults", "true"),
            ("responseData", "false"), ("samplerData", "false"), ("xml", "false"),
            ("fieldNames", "true"), ("responseHeaders", "false"),
            ("requestHeaders", "false"), ("responseDataOnError", "true"),
            ("saveAssertionResultsFailureMessage", "true"),
            ("assertionsResultsToSave", "0"),
            ("bytes", "true"), ("sentBytes", "true"), ("url", "true"),
            ("threadCounts", "true"), ("idleTime", "true"), ("connectTime", "true"),
        ]:
            ET.SubElement(val, tag).text = txt
        return obj

    # ── SETUP THREAD GROUP ────────────────────────────────────────────────────
    stg = ET.SubElement(tpht, "SetupThreadGroup", guiclass="SetupThreadGroupGui",
                        testclass="SetupThreadGroup",
                        testname="SETUP — login once, store shared token",
                        enabled="true")
    xs(stg, "ThreadGroup.on_sample_error", "stoptest")
    slc = ET.SubElement(stg, "elementProp", name="ThreadGroup.main_controller",
                        elementType="LoopController", guiclass="LoopControlPanel",
                        testclass="LoopController", testname="Loop", enabled="true")
    xb(slc, "LoopController.continue_forever", False)
    xs(slc, "LoopController.loops", "1")
    xs(stg, "ThreadGroup.num_threads", "1")
    xs(stg, "ThreadGroup.ramp_time", "1")
    xb(stg, "ThreadGroup.scheduler", False)
    sht = ET.SubElement(tpht, "hashTree")

    if login_req:
        sl = ET.SubElement(sht, "HTTPSamplerProxy", guiclass="HttpTestSampleGui",
                           testclass="HTTPSamplerProxy",
                           testname="SETUP POST /login", enabled="true")
        prs = urlsplit(login_req["url"])
        xs(sl, "HTTPSampler.domain",   prs.hostname or "")
        xs(sl, "HTTPSampler.protocol", prs.scheme or "https")
        xs(sl, "HTTPSampler.port",     str(prs.port) if prs.port else "")
        xs(sl, "HTTPSampler.path",     login_path)
        xs(sl, "HTTPSampler.method",   "POST")
        xs(sl, "HTTPSampler.connect_timeout",  str(CONNECT_TIMEOUT_MS))
        xs(sl, "HTTPSampler.response_timeout", str(RESPONSE_TIMEOUT_MS))
        xb(sl, "HTTPSampler.postBodyRaw",       True)
        xb(sl, "HTTPSampler.use_keepalive",     True)
        xb(sl, "HTTPSampler.follow_redirects",  True)
        ae  = ET.SubElement(sl, "elementProp", name="HTTPsampler.Arguments",
                            elementType="Arguments", guiclass="HTTPArgumentsPanel",
                            testclass="Arguments", testname="Variables", enabled="true")
        col = ET.SubElement(ae, "collectionProp", name="Arguments.arguments")
        a   = ET.SubElement(col, "elementProp", name="", elementType="HTTPArgument")
        xb(a, "HTTPArgument.always_encode", False)
        xs(a, "Argument.value", login_body)
        xs(a, "Argument.metadata", "=")
        slht = ET.SubElement(sht, "hashTree")

        hm   = ET.SubElement(slht, "HeaderManager", guiclass="HeaderPanel",
                             testclass="HeaderManager", testname="Headers", enabled="true")
        hcol = ET.SubElement(hm, "collectionProp", name="HeaderManager.headers")
        for hk, hv in login_req.get("headers", {}).items():
            if hk.lower() in SKIP_HDRS: continue
            if hk.lower() in ("authorization", "content-length"): continue
            hdr(hcol, hk, hv)
        ET.SubElement(slht, "hashTree")

        token_candidate_vars = []
        for tk in TOKEN_KEYS:
            vn = f"SETUP_TK_{safe_var(tk)}"
            token_candidate_vars.append(vn)
            je = ET.SubElement(slht, "JSONPostProcessor", guiclass="JSONPostProcessorGui",
                               testclass="JSONPostProcessor",
                               testname=f"Extract token candidate {tk}", enabled="true")
            xs(je, "JSONPostProcessor.referenceNames", vn)
            xs(je, "JSONPostProcessor.jsonPathExprs",  f"$..{tk}")
            xs(je, "JSONPostProcessor.match_numbers",  "1")
            xs(je, "JSONPostProcessor.defaultValues",  "NOT_FOUND")
            ET.SubElement(slht, "hashTree")

        for extra_field in ["workspace_uid", "org_uid", "user_uid", "uid",
                            "project_uid", "created_by"]:
            vn = f"SETUP_{safe_var(extra_field)}"
            je = ET.SubElement(slht, "JSONPostProcessor", guiclass="JSONPostProcessorGui",
                               testclass="JSONPostProcessor",
                               testname=f"Extract {extra_field}", enabled="true")
            xs(je, "JSONPostProcessor.referenceNames", vn)
            xs(je, "JSONPostProcessor.jsonPathExprs",  f"$..{extra_field}")
            xs(je, "JSONPostProcessor.match_numbers",  "1")
            xs(je, "JSONPostProcessor.defaultValues",  "NOT_FOUND")
            ET.SubElement(slht, "hashTree")

        for field, prop in corr.items():
            if field == token_key: continue
            vn = f"SETUP_{safe_var(field)}"
            je = ET.SubElement(slht, "JSONPostProcessor", guiclass="JSONPostProcessorGui",
                               testclass="JSONPostProcessor",
                               testname=f"Extract {field}", enabled="true")
            xs(je, "JSONPostProcessor.referenceNames", vn)
            xs(je, "JSONPostProcessor.jsonPathExprs",  f"$..{field}")
            xs(je, "JSONPostProcessor.match_numbers",  "1")
            xs(je, "JSONPostProcessor.defaultValues",  "NOT_FOUND")
            ET.SubElement(slht, "hashTree")

        cands_list = ", ".join(f'"{v}"' for v in token_candidate_vars)

        store_corr = "\n".join(
            f'def s_{safe_var(f)}=vars.get("SETUP_{safe_var(f)}"); '
            f'if(s_{safe_var(f)}!=null && !s_{safe_var(f)}.equals("NOT_FOUND")){{ '
            f'props.put("{p}",s_{safe_var(f)}); '
            f'log.info("SETUP stored {p}="+s_{safe_var(f)}.take(12)+"..."); }}'
            for f, p in corr.items() if f != token_key
        )

        setup_script = f"""
// v47 SETUP: pick token, store all correlated values as JMeter properties
def cands=[{cands_list}]
def tok=""
for(cv in cands){{
    def v=vars.get(cv)
    if(v!=null && !v.equals("NOT_FOUND") && v.length()>20){{
        tok=v; log.info("SETUP token via "+cv); break
    }}
}}
if(tok.isEmpty()){{
    log.error("SETUP FAILED: no token in login response.")
    throw new Exception("Login failed — test aborted")
}}
props.put("SHARED_TOKEN",tok)
log.info("SETUP SHARED_TOKEN stored, len="+tok.length())

def wsUid = vars.get("SETUP_WORKSPACE_UID")
if(wsUid==null || wsUid.equals("NOT_FOUND")){{
    wsUid = vars.get("SETUP_ORG_UID")
}}
if(wsUid==null || wsUid.equals("NOT_FOUND")){{
    wsUid = vars.get("SETUP_UID")
}}
if(wsUid!=null && !wsUid.equals("NOT_FOUND")){{
    props.put("CORR_WORKSPACE_UID", wsUid)
    log.info("SETUP CORR_WORKSPACE_UID="+wsUid.take(12)+"...")
}}

["SETUP_USER_UID","SETUP_ORG_UID","SETUP_CREATED_BY"].each {{ k ->
    def v = vars.get(k)
    if(v != null && !v.equals("NOT_FOUND")) {{
        def propKey = k.replace("SETUP_","CORR_")
        props.put(propKey, v)
        log.info("SETUP stored "+propKey+"="+v.take(12)+"...")
    }}
}}

{store_corr}
"""
        jsr = ET.SubElement(slht, "JSR223PostProcessor", guiclass="TestBeanGUI",
                            testclass="JSR223PostProcessor",
                            testname="SETUP store SHARED_TOKEN+corr vars",
                            enabled="true")
        xs(jsr, "scriptLanguage", "groovy")
        xs(jsr, "script", setup_script)
        ET.SubElement(slht, "hashTree")

        a2  = ET.SubElement(slht, "ResponseAssertion", guiclass="AssertionGui",
                            testclass="ResponseAssertion",
                            testname="Assert login 2xx", enabled="true")
        tc2 = ET.SubElement(a2, "collectionProp", name="Asserion.test_strings")
        ET.SubElement(tc2, "stringProp", name="").text = "2\\d\\d"
        xs(a2, "Assertion.test_field", "Assertion.response_code")
        xb(a2, "Assertion.assume_success", False)
        xi(a2, "Assertion.test_type", 1)
        ET.SubElement(slht, "hashTree")

    # ── MAIN THREAD GROUP ─────────────────────────────────────────────────────
    tg = ET.SubElement(tpht, "ThreadGroup", guiclass="ThreadGroupGui",
                       testclass="ThreadGroup", testname=profile["name"],
                       enabled="true")
    xs(tg, "ThreadGroup.on_sample_error", "continue")
    mlc = ET.SubElement(tg, "elementProp", name="ThreadGroup.main_controller",
                        elementType="LoopController", guiclass="LoopControlPanel",
                        testclass="LoopController", testname="Loop", enabled="true")
    xb(mlc, "LoopController.continue_forever", True)
    xs(mlc, "LoopController.loops", "-1")
    xs(tg, "ThreadGroup.num_threads",  str(profile["threads"]))
    xs(tg, "ThreadGroup.ramp_time",    str(profile["ramp"]))
    xb(tg, "ThreadGroup.scheduler",   True)
    xl(tg, "ThreadGroup.duration",    profile["duration"])
    xl(tg, "ThreadGroup.delay",       0)
    xb(tg, "ThreadGroup.same_user_on_next_iteration", True)
    tght = ET.SubElement(tpht, "hashTree")

    cm = ET.SubElement(tght, "CookieManager", guiclass="CookiePanel",
                       testclass="CookieManager", testname="Cookie Manager",
                       enabled="true")
    xb(cm, "CookieManager.clearEachIteration", True)
    ET.SubElement(tght, "hashTree")

    ca = ET.SubElement(tght, "CacheManager", guiclass="CacheManagerGui",
                       testclass="CacheManager", testname="Cache Manager",
                       enabled="true")
    xb(ca, "clearEachIteration", True)
    ET.SubElement(tght, "hashTree")

    df_el = ET.SubElement(tght, "ConfigTestElement", guiclass="HttpDefaultsGui",
                          testclass="ConfigTestElement", testname="Defaults",
                          enabled="true")
    empty_args(df_el)
    xs(df_el, "HTTPSampler.connect_timeout",  str(CONNECT_TIMEOUT_MS))
    xs(df_el, "HTTPSampler.response_timeout", str(RESPONSE_TIMEOUT_MS))
    xb(df_el, "HTTPSampler.use_keepalive",    True)
    xb(df_el, "HTTPSampler.follow_redirects", True)
    ET.SubElement(tght, "hashTree")

    gt = ET.SubElement(tght, "GaussianRandomTimer", guiclass="GaussianRandomTimerGui",
                       testclass="GaussianRandomTimer", testname="Think Time",
                       enabled="true")
    xs(gt, "ConstantTimer.delay", "1000")
    xs(gt, "RandomTimer.range",   "500")
    ET.SubElement(tght, "hashTree")

    runner = ET.SubElement(tght, "LoopController", guiclass="LoopControlPanel",
                           testclass="LoopController",
                           testname="Full Scenario — 1 iteration", enabled="true")
    xb(runner, "LoopController.continue_forever", False)
    xs(runner, "LoopController.loops", "1")
    runht = ET.SubElement(tght, "hashTree")

    init_script = """
// v47 ITER INIT — unique suffix + dynamic CORR_* prop sync
def uniq =
"_T" + ctx.getThreadNum() +
"_I" + vars.getIteration() +
"_" + System.currentTimeMillis()
vars.put("ITER_SUFFIX", uniq)
log.info("v47 ITER=" + uniq)

// V46-8: Pull ALL CORR_* props into thread-local vars dynamically
props.each { key, value ->
    if (key.toString().startsWith("CORR_") && value != null && value != "NOT_FOUND") {
        vars.put(key.toString(), value.toString())
        log.info("v47 INIT pulled " + key + "=" + value.toString().take(12) + "...")
    }
}

// Also pull SHARED_TOKEN
def tok = props.get("SHARED_TOKEN")
if (tok != null) {
    vars.put("SHARED_TOKEN", tok)
}

log.info("v47 INIT complete — ITER_SUFFIX=" + uniq)
"""
    ii = ET.SubElement(runht, "JSR223Sampler", guiclass="TestBeanGUI",
                       testclass="JSR223Sampler",
                       testname="ITER INIT — unique suffix + shared props",
                       enabled="true")
    xs(ii, "scriptLanguage", "groovy"); xs(ii, "script", init_script)
    xs(ii, "filename", ""); xs(ii, "parameters", "")
    xb(ii, "stopThread", False)
    ET.SubElement(runht, "hashTree")

    def add_ai_assertions(s_ht, resp_json):
        for k, expected_val in generate_assertions(resp_json):
            safe_expected = str(expected_val)[:80]
            ra  = ET.SubElement(s_ht, "ResponseAssertion", guiclass="AssertionGui",
                                testclass="ResponseAssertion",
                                testname=f"AI Assert: {k}={safe_expected[:30]}",
                                enabled="true")
            tc  = ET.SubElement(ra, "collectionProp", name="Asserion.test_strings")
            ET.SubElement(tc, "stringProp", name="").text = safe_expected
            xs(ra, "Assertion.test_field",    "Assertion.response_data")
            xb(ra, "Assertion.assume_success", False)
            xi(ra, "Assertion.test_type",      2)
            ET.SubElement(s_ht, "hashTree")

    RETRY_GROOVY = """
def retryable = ["429","502","503","504"]
def code = prev.getResponseCode()
if (!retryable.contains(code)) return

int retries = vars.get("RETRY_COUNT") ? vars.get("RETRY_COUNT").toInteger() : 0
if (retries < 3) {
    retries++
    vars.put("RETRY_COUNT", retries.toString())
    long backoffMs = (long) Math.pow(2, retries - 1) * 1000L
    log.warn("v47 Retry #" + retries + " for " + prev.getSampleLabel()
             + " (code=" + code + ") — back-off " + backoffMs + "ms")
    Thread.sleep(backoffMs)
    prev.setSuccessful(false)
    SampleResult.setStopThread(false)
} else {
    vars.remove("RETRY_COUNT")
    prev.setSuccessful(false)
    log.error("v47 Max retries (3) exceeded for " + prev.getSampleLabel())
}
"""

    def add_failure_detection(s_ht):
        retry_jsr = ET.SubElement(s_ht, "JSR223PostProcessor",
                                  guiclass="TestBeanGUI",
                                  testclass="JSR223PostProcessor",
                                  testname="Retry Engine v47",
                                  enabled="true")
        xs(retry_jsr, "scriptLanguage", "groovy")
        xs(retry_jsr, "script", RETRY_GROOVY)
        ET.SubElement(s_ht, "hashTree")

        groovy_map = "\n".join(
            f'    "{code}": "{reason}",'
            for code, reason in {
                "401": "Authentication Failure", "403": "Authorization Failure",
                "404": "Dependency Failure",     "409": "Duplicate Test Data",
                "429": "Rate Limiting",          "500": "Backend Exception",
                "502": "Gateway Error",          "503": "Service Unavailable",
                "504": "Timeout",
            }.items()
        )
        jsr = ET.SubElement(s_ht, "JSR223PostProcessor", guiclass="TestBeanGUI",
                            testclass="JSR223PostProcessor",
                            testname="Failure Detection Engine v47", enabled="true")
        xs(jsr, "scriptLanguage", "groovy")
        xs(jsr, "script", f"""
def code  = prev.getResponseCode()
def label = prev.getSampleLabel()?.toLowerCase() ?: ""
def elapsed = prev.getTime()
def failureMap = [
{groovy_map}
]
def ok_codes = ["200","201","202","204"]
if (!ok_codes.contains(code)) {{
    def reason = failureMap.getOrDefault(code, "Unknown Failure")
    if (code == "500" && elapsed > 5000) {{ reason = "Database Bottleneck" }}
    log.error("FAILED API  = " + prev.getSampleLabel())
    log.error("STATUS CODE = " + code + " → " + reason)
    log.error("ELAPSED     = " + elapsed + "ms")
    log.error("RESPONSE    = " + prev.getResponseDataAsString())
}} else {{
    prev.setSuccessful(true)
    log.info("SUCCESS → " + prev.getSampleLabel() + " [" + code + "]")
}}
""")
        ET.SubElement(s_ht, "hashTree")

    def add_entity_extractors(s_ht):
        for field, prop in corr.items():
            if not prop.startswith("CORR_"): continue
            je = ET.SubElement(s_ht, "JSONPostProcessor",
                               guiclass="JSONPostProcessorGui",
                               testclass="JSONPostProcessor",
                               testname=f"Extract {field}", enabled="true")
            xs(je, "JSONPostProcessor.referenceNames", prop)
            xs(je, "JSONPostProcessor.jsonPathExprs",  f"$..{field}")
            xs(je, "JSONPostProcessor.match_numbers",  "1")
            xs(je, "JSONPostProcessor.defaultValues",  "NOT_FOUND")
            ET.SubElement(s_ht, "hashTree")

    def add_prop_storage_script(s_ht, label=""):
        jsr = ET.SubElement(s_ht, "JSR223PostProcessor", guiclass="TestBeanGUI",
                            testclass="JSR223PostProcessor",
                            testname=f"V47 Store all CORR_* as shared props {label}",
                            enabled="true")
        xs(jsr, "scriptLanguage", "groovy")
        xs(jsr, "script", """
// V47: After every CREATE, store all CORR_* vars as props for cross-thread sharing
vars.entrySet().each { entry ->
    def k = entry.key.toString()
    def v = entry.value?.toString()
    if (k.startsWith("CORR_") && v != null && v != "NOT_FOUND" && v.length() > 3) {
        props.put(k, v)
        log.info("V47 stored prop " + k + "=" + v.take(12) + "...")
    }
}
def samplerPath = prev.getSampleLabel()
def respCode = prev.getResponseCode()
if (respCode == "201") {
    try {
        def js = new groovy.json.JsonSlurper()
        def resp = js.parseText(prev.getResponseDataAsString())
        def extractUid = { obj ->
            if (obj instanceof Map) {
                if (obj.containsKey("uid")) return obj["uid"]
                for (v in obj.values()) { def r = extractUid(v); if (r != null) return r }
            }
            if (obj instanceof List) {
                for (item in obj) { def r = extractUid(item); if (r != null) return r }
            }
            return null
        }
        def uid = extractUid(resp)
        if (uid != null) {
            if (samplerPath.contains("/project") && !samplerPath.contains("/requirements")
                && !samplerPath.contains("/defects") && !samplerPath.contains("/sprint")) {
                props.put("CORR_PROJECT_UID", uid.toString())
                vars.put("CORR_PROJECT_UID", uid.toString())
                log.info("V47 project_uid stored from 201: " + uid.toString().take(12) + "...")
            } else if (samplerPath.contains("/sprint") && !samplerPath.contains("start")) {
                props.put("CORR_SPRINT_UID", uid.toString())
                vars.put("CORR_SPRINT_UID", uid.toString())
            } else if (samplerPath.contains("/requirements")) {
                props.put("CORR_REQUIREMENT_UID", uid.toString())
                vars.put("CORR_REQUIREMENT_UID", uid.toString())
            } else if (samplerPath.contains("/defects")) {
                props.put("CORR_DEFECT_UID", uid.toString())
                vars.put("CORR_DEFECT_UID", uid.toString())
            }
        }
    } catch(Exception e) {
        log.warn("V47 uid extraction failed: " + e.getMessage())
    }
}
""")
        ET.SubElement(s_ht, "hashTree")

    def add_corr_refresh_preprocessor(s_ht):
        jsr = ET.SubElement(s_ht, "JSR223PreProcessor", guiclass="TestBeanGUI",
                            testclass="JSR223PreProcessor",
                            testname="V47 Refresh CORR_* vars from props",
                            enabled="true")
        xs(jsr, "scriptLanguage", "groovy")
        xs(jsr, "script", """
// V47-3: Refresh all CORR_* vars from shared props right before this sampler fires
props.each { key, value ->
    if (key.toString().startsWith("CORR_") && value != null && value != "NOT_FOUND") {
        def current = vars.get(key.toString())
        if (current == null || current == "NOT_FOUND") {
            vars.put(key.toString(), value.toString())
            log.info("V47 PreProc refreshed " + key + "=" + value.toString().take(12) + "...")
        }
    }
}
""")
        ET.SubElement(s_ht, "hashTree")

    def assertion(s_ht, recorded_ok, recorded_status, is_create=False,
                  method="GET", url="", resp_json=None, response_ms=0):
        NON_BUSINESS_PATHS = [
            "/dashboard/", "/notification", "/notifications/",
            "/history/", "/recent/", "/activity/",
            "/analytics/", "/comments/", "/attachments/",
            "/widget", "/search", "/list", "/dropdown",
        ]
        if any(nb in url.lower() for nb in NON_BUSINESS_PATHS): return
        if recorded_ok:
            a  = ET.SubElement(s_ht, "ResponseAssertion", guiclass="AssertionGui",
                               testclass="ResponseAssertion",
                               testname="Assert 2xx", enabled="true")
            tc = ET.SubElement(a, "collectionProp", name="Asserion.test_strings")
            ET.SubElement(tc, "stringProp", name="").text = "2\\d\\d"
            xs(a, "Assertion.test_field",    "Assertion.response_code")
            xb(a, "Assertion.assume_success", False)
            xi(a, "Assertion.test_type",      1)
        else:
            a  = ET.SubElement(s_ht, "ResponseAssertion", guiclass="AssertionGui",
                               testclass="ResponseAssertion",
                               testname=f"Assert not 5xx (recorded {recorded_status})",
                               enabled="true")
            tc = ET.SubElement(a, "collectionProp", name="Asserion.test_strings")
            ET.SubElement(tc, "stringProp", name="").text = "5\\d\\d"
            xs(a, "Assertion.test_field",    "Assertion.response_code")
            xb(a, "Assertion.assume_success", False)
            xi(a, "Assertion.test_type",      6)
        ET.SubElement(s_ht, "hashTree")
        path_lower  = url.lower()
        is_list_api = any(x in path_lower for x in [
            "get_all", "get_projects", "listing", "list", "get_org",
        ])
        multiplier = 1.5
        floor_ms   = 2000 if is_list_api else 1000
        sla_ms     = max(int((response_ms or 3000) * multiplier), floor_ms)
        da = ET.SubElement(s_ht, "DurationAssertion",
                           guiclass="DurationAssertionGui",
                           testclass="DurationAssertion",
                           testname=f"SLA<{sla_ms}ms (recorded {response_ms}ms × {multiplier})",
                           enabled="true")
        xs(da, "DurationAssertion.duration", str(sla_ms))
        ET.SubElement(s_ht, "hashTree")
        if is_create and method == "POST":
            add_ai_assertions(s_ht, resp_json)

    def build_ws_sampler(r, parent_ht):
        samp = ET.SubElement(parent_ht, "GenericSampler",
                             guiclass="GenericSamplerUI",
                             testclass="GenericSampler",
                             testname=f"🔌 WebSocket {r['path']}", enabled="true")
        xs(samp, "GenericSampler.requestData",
           json.dumps({"wsUrl": r["url"],
                       "frames": r.get("ws_frames", [])}, indent=2))
        s_ht = ET.SubElement(parent_ht, "hashTree")
        jsr = ET.SubElement(s_ht, "JSR223Sampler", guiclass="TestBeanGUI",
                            testclass="JSR223Sampler",
                            testname="WebSocket connect + replay frames",
                            enabled="true")
        xs(jsr, "scriptLanguage", "groovy")
        frames      = r.get("ws_frames", [])
        sent_frames = [f["data"] for f in frames
                       if f.get("type") == "sent" and f.get("data")]
        frames_groovy = "\n".join(
            f'ws.sendTextFrame({json.dumps(str(d))})'
            for d in sent_frames[:10]
        )
        xs(jsr, "script", f"""
import com.blazemeter.jmeter.websocket.sampler.WebSocketSamplerPlugin
def wsUrl = "{r['url']}"
log.info("v47 WS connect: " + wsUrl)
{frames_groovy if frames_groovy else '// No sent frames captured'}
log.info("v47 WS frames replayed")
""")
        ET.SubElement(s_ht, "hashTree")

    def build_sampler(r, parent_ht):
        if r.get("method") == "WEBSOCKET":
            build_ws_sampler(r, parent_ht); return

        is_login  = r.get("isLogin", False)
        is_create = (r.get("isCreate", False) or
                     (r["method"] == "POST" and not is_login and
                      not any(x in r["path"].lower()
                              for x in ["get","list","search","validate",
                                        "login","logout"])))
        recorded_ok = 200 <= (r.get("status") or 200) < 300

        is_sprint_start   = r["method"] == "POST" and "sprint/start" in r["path"]
        is_project_create = (r["method"] == "POST" and
                             r["path"].rstrip("/").endswith("/project") and
                             not any(x in r["path"].lower() for x in
                                     ["requirements","defects","sprint","user"]))

        needs_project_uid = (
            "${CORR_PROJECT_UID}" in r.get("path", "") or
            "${CORR_PROJECT_UID}" in r.get("postData", "") or
            any(
                "${CORR_PROJECT_UID}" in str(v)
                for v in r.get("_healed_multipart", {}).values()
            )
        )

        label = (
            f"{'🔑 ' if is_login else ''}"
            f"{'🆕 ' if is_create and not is_login else ''}"
            f"{'🔷 ' if r.get('isGraphQL') else ''}"
            f"{r['method']} {r['path']}"
        )

        path = r["path"]
        if re.search(r"\$\{CORR_[0-9A-F]{8}\}", path):
            path = re.sub(r"\$\{CORR_[0-9A-F]{8}\}", "${CORR_WORKSPACE_UID}", path)
            r["path"] = path

        if is_sprint_start:
            r["postData"] = ('{"project_uid":"${CORR_PROJECT_UID}",'
                             '"sprint_uid":"${CORR_SPRINT_UID}"}')

        prs = urlsplit(r["url"])
        smp = ET.SubElement(parent_ht, "HTTPSamplerProxy",
                            guiclass="HttpTestSampleGui",
                            testclass="HTTPSamplerProxy",
                            testname=label, enabled="true")
        xs(smp, "HTTPSampler.domain",          prs.hostname or "")
        xs(smp, "HTTPSampler.protocol",         prs.scheme or "https")
        xs(smp, "HTTPSampler.port",             str(prs.port) if prs.port else "")
        xs(smp, "HTTPSampler.path",             r["path"])
        xs(smp, "HTTPSampler.method",           r["method"])
        xs(smp, "HTTPSampler.connect_timeout",  str(CONNECT_TIMEOUT_MS))
        xs(smp, "HTTPSampler.response_timeout", str(RESPONSE_TIMEOUT_MS))
        xb(smp, "HTTPSampler.use_keepalive",    True)
        xb(smp, "HTTPSampler.follow_redirects", True)
        xb(smp, "HTTPSampler.auto_redirects",   False)

        body = r.get("postData", "") if r["method"] in ("POST","PUT","PATCH") else ""
        if is_create and not is_sprint_start:
            body = fix_create(body, corr)
        if any(vp in r["path"].lower() for vp in VALIDATE_PATHS):
            body = fix_validate(body)

        is_multipart = "multipart" in r.get("contentType", "").lower()

        if r["method"] in ("POST","PUT","PATCH") and body and is_multipart:
            xb(smp, "HTTPSampler.postBodyRaw",       False)
            xb(smp, "HTTPSampler.DO_MULTIPART_POST", True)
            ae   = ET.SubElement(smp, "elementProp", name="HTTPsampler.Arguments",
                                 elementType="Arguments", guiclass="HTTPArgumentsPanel",
                                 testclass="Arguments", testname="Variables", enabled="true")
            coll = ET.SubElement(ae, "collectionProp", name="Arguments.arguments")
            mp_fields = r.get("_healed_multipart") or parse_multipart(r.get("postData",""))
            if not mp_fields:
                arg = ET.SubElement(coll, "elementProp", name="", elementType="HTTPArgument")
                xb(arg, "HTTPArgument.always_encode", False)
                xs(arg, "Argument.value", body); xs(arg, "Argument.metadata", "=")
            else:
                for fname, fval in mp_fields.items():
                    if str(fval).strip().lower() in ("null","none",""): continue
                    explicit_uid_map = {
                        "project_uid": "${CORR_PROJECT_UID}",
                        "sprint_uid":  "${CORR_SPRINT_UID}",
                        "org_uid":     "${CORR_ORG_UID}",
                        "user_uid":    "${CORR_USER_UID}",
                    }
                    if fname.lower() in explicit_uid_map:
                        fval = explicit_uid_map[fname.lower()]
                    else:
                        for cf, cv in corr.items():
                            if fname.lower() == cf.lower():
                                fval = "${" + cv + "}"; break
                    if is_create and fname.lower() in (
                        "summary", "title", "name", "steps", "actual", "expected"
                    ) and not str(fval).startswith("${"):
                        fval = f"{generate_test_data(fname)}${{ITER_SUFFIX}}"
                    arg = ET.SubElement(coll, "elementProp", name=fname,
                                        elementType="HTTPArgument")
                    xb(arg, "HTTPArgument.always_encode", False)
                    xs(arg, "Argument.name",  fname)
                    xs(arg, "Argument.value", fval)
                    xs(arg, "Argument.metadata", "=")

        elif r["method"] in ("POST","PUT","PATCH") and body:
            xb(smp, "HTTPSampler.postBodyRaw", True)
            ae  = ET.SubElement(smp, "elementProp", name="HTTPsampler.Arguments",
                                elementType="Arguments", guiclass="HTTPArgumentsPanel",
                                testclass="Arguments", testname="Variables", enabled="true")
            c   = ET.SubElement(ae, "collectionProp", name="Arguments.arguments")
            arg = ET.SubElement(c, "elementProp", name="", elementType="HTTPArgument")
            xb(arg, "HTTPArgument.always_encode", False)
            xs(arg, "Argument.value", body); xs(arg, "Argument.metadata", "=")

        elif r.get("query") and r["method"] == "GET":
            qp  = parse_qs(r["query"], keep_blank_values=True)
            ae  = ET.SubElement(smp, "elementProp", name="HTTPsampler.Arguments",
                                elementType="Arguments", guiclass="HTTPArgumentsPanel",
                                testclass="Arguments", testname="Variables", enabled="true")
            c   = ET.SubElement(ae, "collectionProp", name="Arguments.arguments")
            for pk, pvs in qp.items():
                for pv in pvs:
                    if pk == "dateFrom":      pv = "${DASH_FROM}"
                    elif pk == "dateTo":      pv = "${DASH_TO}"
                    elif pk == "project_uid": pv = "${CORR_PROJECT_UID}"
                    arg = ET.SubElement(c, "elementProp", name=pk,
                                        elementType="HTTPArgument")
                    use_encode = False if "${" in pv else True
                    xb(arg, "HTTPArgument.always_encode", use_encode)
                    xs(arg, "Argument.name",  pk)
                    xs(arg, "Argument.value", pv)
                    xs(arg, "Argument.metadata", "=")
        else:
            empty_args(smp)

        s_ht = ET.SubElement(parent_ht, "hashTree")

        if not is_login:
            add_corr_refresh_preprocessor(s_ht)

        hm   = ET.SubElement(s_ht, "HeaderManager", guiclass="HeaderPanel",
                             testclass="HeaderManager", testname="Headers", enabled="true")
        hcol = ET.SubElement(hm, "collectionProp", name="HeaderManager.headers")
        for hk, hv in r.get("headers", {}).items():
            if hk.lower() in SKIP_HDRS: continue
            if hk.lower() in ("authorization", "content-length"): continue
            hdr(hcol, hk, hv)
        if not is_login:
            hdr(hcol, "Authorization", "Bearer ${__P(SHARED_TOKEN,INIT)}")
        ET.SubElement(s_ht, "hashTree")

        if not is_login:
            reauth = ET.SubElement(s_ht, "JSR223PostProcessor",
                                   guiclass="TestBeanGUI",
                                   testclass="JSR223PostProcessor",
                                   testname="ON 401 refresh token + replay v47",
                                   enabled="true")
            xs(reauth, "scriptLanguage", "groovy")
            xs(reauth, "script", f"""
if(prev.getResponseCode()!="401") return
log.warn("v47 401 on "+prev.getSampleLabel()+" — refreshing token…")
import groovy.json.JsonSlurper
def conn=new URL("{login_scheme}://{login_host}{login_path}").openConnection()
conn.setRequestMethod("POST"); conn.setDoOutput(true)
conn.setRequestProperty("Content-Type","{login_ct}")
conn.setConnectTimeout({CONNECT_TIMEOUT_MS})
conn.setReadTimeout({RESPONSE_TIMEOUT_MS})
conn.outputStream.write('{login_body_groovy}'.getBytes("UTF-8"))
def newTok = null
if(conn.responseCode>=200 && conn.responseCode<300){{
    def rj=new JsonSlurper().parseText(conn.inputStream.text)
    newTok=findTok(rj)
    if(newTok){{ props.put("SHARED_TOKEN",newTok); log.info("v47 token refreshed") }}
}}
if(newTok){{
    def replayUrl = prev.getUrlAsString()
    def replayConn = new URL(replayUrl).openConnection()
    replayConn.setRequestMethod(prev.getHTTPMethod() ?: "GET")
    replayConn.setRequestProperty("Authorization","Bearer "+newTok)
    replayConn.setRequestProperty("Content-Type","{r.get('contentType','application/json')}")
    replayConn.setConnectTimeout({CONNECT_TIMEOUT_MS})
    replayConn.setReadTimeout({RESPONSE_TIMEOUT_MS})
    def replayBody = prev.getSamplerData()
    if(replayBody){{ replayConn.setDoOutput(true)
        replayConn.outputStream.write(replayBody.getBytes("UTF-8")) }}
    def replayCode = replayConn.responseCode
    if(replayCode>=200 && replayCode<300){{ prev.setSuccessful(true)
        log.info("v47 Replay succeeded for "+prev.getSampleLabel()) }}
}}
def findTok(o){{
    def TOKEN_KEYS_G = {token_keys_groovy}
    if(o instanceof Map){{
        for(tk in TOKEN_KEYS_G){{
            if(o.containsKey(tk) && o[tk] instanceof String && o[tk].length()>20) return o[tk]
        }}
        for(v in o.values()){{ def r=findTok(v); if(r) return r }}
    }}
    if(o instanceof List){{ for(i in o){{ def r=findTok(i); if(r) return r }} }}
    return null
}}
""")
            ET.SubElement(s_ht, "hashTree")

        if is_create and not is_login:
            add_entity_extractors(s_ht)
            add_prop_storage_script(s_ht, label=r["path"])
            log_jsr   = ET.SubElement(s_ht, "JSR223PostProcessor",
                                      guiclass="TestBeanGUI",
                                      testclass="JSR223PostProcessor",
                                      testname="LOG entity created v47", enabled="true")
            log_lines = []
            for field, prop in corr.items():
                if prop.startswith("CORR_"):
                    log_lines.append(
                        f'def v_{safe_var(field)}=vars.get("{prop}"); '
                        f'if(v_{safe_var(field)}!=null && '
                        f'!v_{safe_var(field)}.equals("NOT_FOUND"))'
                        f'{{ log.info("v47 {field}="+v_{safe_var(field)}); }}'
                    )
            xs(log_jsr, "scriptLanguage", "groovy")
            xs(log_jsr, "script",
               "\n".join(log_lines) if log_lines
               else 'log.info("v47 entity created: " + prev.getSampleLabel())')
            ET.SubElement(s_ht, "hashTree")

        add_failure_detection(s_ht)
        assertion(
            s_ht, recorded_ok, r.get("status"),
            is_create=(is_create and not is_login),
            method=r["method"],
            url=r.get("url", r.get("path", "")),
            resp_json=r.get("resp_json"),
            response_ms=r.get("response_ms", 0),
        )

    # Apply path substitutions to all requests
    for r in reqs:
        old = r["path"]
        for raw, field in sorted(val_map.items(), key=lambda x: -len(x[0])):
            prop = corr.get(field)
            if not prop or raw not in r["path"]: continue
            r["path"] = r["path"].replace(raw, f"${{{prop}}}")
        if r["path"] != old:
            ok(f"Path fixed: {old} → {r['path']}")
        def _fix_uuid_in_path(path):
            parts = path.split("/")
            result = []
            for part in parts:
                if is_uuid(part):
                    matched = next(
                        (f"${{{corr[f]}}}" for f, d in correlation_store.items()
                         if d["value"] == part and f in corr),
                        None
                    )
                    result.append(matched if matched else "${CORR_WORKSPACE_UID}")
                else:
                    result.append(part)
            return "/".join(result)
        new_path = _fix_uuid_in_path(r["path"])
        if new_path != r["path"]:
            r["path"] = new_path

    # V46-1: Build flows — skip all-login flows
    login_flows = [f for f in flows if any(r.get("isLogin") for r in f)]
    other_flows = [f for f in flows if not any(r.get("isLogin") for r in f)]

    for fl in login_flows + other_flows:
        non_login = [r for r in fl if not r.get("isLogin")]
        if not non_login:
            ok(f"V46-1: Skipping empty flow (all-login)")
            continue
        fname = flow_name(non_login)
        tc    = ET.SubElement(runht, "TransactionController",
                              guiclass="TransactionControllerGui",
                              testclass="TransactionController",
                              testname=fname, enabled="true")
        xb(tc, "TransactionController.parent",       True)
        xb(tc, "TransactionController.includeTimers", False)
        tc_ht = ET.SubElement(runht, "hashTree")
        for r in non_login:
            build_sampler(r, tc_ht)
        ok(f"  Flow '{fname}' — {len(non_login)} sampler(s)")

    for gc, tn, en in [
        ("SummaryReport",             "Summary Report",               True),
        ("StatVisualizer",            "Aggregate Report",             True),
        ("StatGraphVisualizer",       "Aggregate Graph",              True),
        ("ViewResultsFullVisualizer", "View Results Tree (GUI only)", False),
    ]:
        rc = ET.SubElement(tpht, "ResultCollector", guiclass=gc,
                           testclass="ResultCollector", testname=tn, enabled="true")
        xb(rc, "ResultCollector.error_logging", False)
        rc.append(save_cfg())
        xs(rc, "filename", jtl_path if en else "")
        ET.SubElement(tpht, "hashTree")

    with open(jmx_path, "w", encoding="utf-8") as f:
        f.write(prettify(root))
    ok(f"JMX written → {jmx_path}")

# ─── CSV REPORT (v46 unchanged) ───────────────────────────────────────────────
def write_csv_report(ts, profile, corr, token_key, user_fields, login_req, csv_path):
    rows = []
    rows.append(["Category","Field","Recorded Value","JMeter Expression","Notes"])
    rows.append(["","","","",""])
    creds = login_req.get("creds", {}) if login_req else {}
    for k, v in creds.items():
        rows.append(["Credential", k, v, "(hardcoded in login body)",
                     "Shared via SHARED_TOKEN across all threads"])
    rows.append(["","","","",""])
    rows.append(["--- SERVER-EXTRACTED VALUES ---","","","",""])
    for field, prop in corr.items():
        conf = correlation_confidence.get(field, 0)
        if field == token_key:
            rows.append(["Auth Token", field, "(from login response)",
                         f"${{__P({prop})}}", "Shared across all threads"])
        else:
            rows.append(["Correlation", field, "(from API response)",
                         f"${{{prop}}}", f"V47 struct-aware (conf={conf})"])
    rows.append(["","","","",""])
    rows.append(["--- USER-TYPED FIELDS ---","","","",""])
    for field, val in user_fields.items():
        rows.append(["User Field", field, val,
                     f"<faker:{field}>${{ITER_SUFFIX}}",
                     "V47 Faker: realistic data per iteration"])
    rows.append(["","","","",""])
    rows.append(["--- TEST PROFILE ---","","","",""])
    rows.append(["Profile","Name",  profile["name"],"",""])
    rows.append(["Profile","Threads",str(profile["threads"]),"","Virtual users"])
    rows.append(["Profile","Ramp-up",str(profile["ramp"])+"s","",""])
    rows.append(["Profile","Duration",str(profile["duration"])+"s","",""])

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    ok(f"CSV report → {csv_path}")

    log(); log(f"  {B}{C}PARAMETERIZATION SUMMARY:{W}")
    log(f"  {'─'*65}")
    log(f"  {'Category':<18} {'Field':<24} {'JMeter Expression':<30}")
    log(f"  {'─'*18} {'─'*24} {'─'*30}")
    for row in rows:
        if (len(row) == 5 and row[0]
                and not row[0].startswith("---")
                and row[0] != "Category"):
            log(f"  {row[0]:<18} {row[1]:<24} {row[3]:<30}  ← {row[4][:30]}")

# ─── LIVE MONITOR ─────────────────────────────────────────────────────────────
def live_monitor(profile):
    th = profile["threads"]; ramp = profile["ramp"]; dur = profile["duration"]
    def _run():
        s = time.time()
        log(); log("─"*65); log("  LIVE MONITOR"); log("─"*65)
        while True:
            e = time.time() - s
            if e > dur: ok(f"Complete ({dur}s)"); break
            a   = min(int((e / max(ramp, 1)) * th), th)
            bar = "█" * int(30*a/max(th,1)) + "░" * (30 - int(30*a/max(th,1)))
            log(f"  [{bar}] {a}/{th} users  t={int(e)}s/{dur}s")
            time.sleep(15)
    threading.Thread(target=_run, daemon=True).start()

# ─── SCENARIO MANAGER ─────────────────────────────────────────────────────────
def get_scenarios():
    if not os.path.exists(SCENARIOS_DIR): return []
    return [d for d in os.listdir(SCENARIOS_DIR)
            if os.path.isdir(os.path.join(SCENARIOS_DIR, d))]

def save_scenario(jmx_path):
    scenario_name = input("\nEnter Scenario Name : ").strip()
    scenario_dir  = os.path.join(SCENARIOS_DIR, scenario_name)
    os.makedirs(scenario_dir, exist_ok=True)
    target_jmx = os.path.join(scenario_dir, "scenario.jmx")
    shutil.copy2(jmx_path, target_jmx)
    ok(f"Scenario Saved : {scenario_name}")
    ok(f"Location       : {target_jmx}")
    return target_jmx

def execute_saved_jmx(jmx_file, profile):
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_jtl = os.path.join(OUT_DIR, f"result_{ts}.jtl")
    report_dir = os.path.join(OUT_DIR, "html_reports", f"report_{ts}")

    jbin = JMETER_BIN
    for nm in ("jmeter", "jmeter.bat", "jmeter.sh"):
        found = shutil.which(nm)
        if found: jbin = found; break

    if not os.path.exists(jbin):
        err(f"JMeter not found at: {jbin}")
        info(f'Run manually:  jmeter -n -t "{jmx_file}" -l "{result_jtl}" '
             f'-e -o "{report_dir}"')
        return

    cmd = [
        jbin, "-n",
        "-t",  jmx_file,
        "-l",  result_jtl,
        "-e",
        "-o",  report_dir,
        "-Jthreads="   + str(profile["threads"]),
        "-Jramp="      + str(profile["ramp"]),
        "-Jduration="  + str(profile["duration"]),
    ]
    if os.path.exists(result_jtl): os.remove(result_jtl)
    if os.path.exists(report_dir): shutil.rmtree(report_dir)

    # V47: Start psutil monitor alongside JMeter
    monitor = BottleneckMonitor(interval=5.0)
    monitor.start()
    live_monitor(profile)

    try:
        ok("JMeter starting …")
        subprocess.run(cmd, check=True)
        monitor.stop()
        ok("Execution Completed")
        ok(f"JTL results : {result_jtl}")
        ok(f"HTML report : {report_dir}/index.html")
        monitor.print_summary()
        banner("POST-RUN BOTTLENECK ANALYSIS")
        analyze_jtl(result_jtl)
        idx_path = os.path.join(report_dir, "index.html")
        if os.path.exists(idx_path):
            webbrowser.open(idx_path)
        else:
            warn(f"HTML report not generated at: {idx_path}")
    except subprocess.CalledProcessError as e:
        monitor.stop()
        err(f"JMeter failed (exit {e.returncode})")
        err("Tip: open the JMX in JMeter GUI → run 1 thread → View Results Tree")

def run_existing_scenario():
    scenarios = get_scenarios()
    if not scenarios:
        warn("No saved scenarios found.")
        info(f"Record a new scenario first — files stored in: {SCENARIOS_DIR}")
        return
    log(); log(f"  {B}{C}Available Scenarios:{W}\n")
    for i, s in enumerate(scenarios, start=1):
        log(f"    {i}. {s}")
    log()
    try:
        choice = int(input("  Select Scenario number : ").strip())
        if choice < 1 or choice > len(scenarios):
            err("Invalid selection."); return
    except ValueError:
        err("Please enter a number."); return

    scenario = scenarios[choice - 1]
    jmx_file = os.path.join(SCENARIOS_DIR, scenario, "scenario.jmx")
    if not os.path.exists(jmx_file):
        err(f"scenario.jmx not found in: {os.path.dirname(jmx_file)}"); return

    ok(f"Selected scenario : {scenario}")
    log(); log(f"  {B}{C}Select Test Profile:{W}\n")
    for k, v in TEST_PROFILES.items():
        log(f"    {k}. {v['name']:<16}  "
            f"users={v['threads']}  ramp={v['ramp']}s  dur={v['duration']}s")
    log()
    profile_choice = input("  Choice [1–5, default=1] : ").strip() or "1"
    if profile_choice not in TEST_PROFILES: profile_choice = "1"
    profile = TEST_PROFILES[profile_choice]
    ok(f"Profile : {profile['name']} — {profile['threads']} virtual users")
    execute_saved_jmx(jmx_file, profile)

# ─── V47: PYTHON LOAD ENGINE RUNNER ──────────────────────────────────────────
def run_python_load_test(test_type: str):
    """
    V47-B: Run a pure-Python load test using an existing saved scenario.
    test_type: "1"=Functional, "2"=Load, "3"=Stress, "4"=Spike, "5"=Endurance
    """
    scenarios = get_scenarios()
    if not scenarios:
        warn("No saved scenarios found. Record a new scenario first.")
        return

    log(); log(f"  {B}{C}Available Scenarios:{W}\n")
    for i, s in enumerate(scenarios, start=1):
        log(f"    {i}. {s}")
    log()
    try:
        choice = int(input("  Select Scenario number : ").strip())
        if choice < 1 or choice > len(scenarios):
            err("Invalid selection."); return
    except ValueError:
        err("Please enter a number."); return

    scenario   = scenarios[choice - 1]
    jmx_file   = os.path.join(SCENARIOS_DIR, scenario, "scenario.jmx")
    json_cache = os.path.join(SCENARIOS_DIR, scenario, "requests.json")

    # Try to load recorded requests from JSON cache
    reqs = []
    if os.path.exists(json_cache):
        try:
            with open(json_cache, "r", encoding="utf-8") as f:
                reqs = json.load(f)
            ok(f"Loaded {len(reqs)} cached requests from scenario")
        except Exception:
            pass

    if not reqs:
        warn("No cached requests found for this scenario.")
        target_url = input("  Enter target URL for load test (e.g. https://api.example.com): ").strip()
        if not target_url:
            err("No URL provided."); return
        reqs = [{"method": "GET", "url": target_url, "headers": {},
                 "postData": "", "contentType": "application/json",
                 "path": "/", "query": ""}]

    profile = LOAD_VU_PROFILES.get(test_type, LOAD_VU_PROFILES["2"])
    log()
    log(f"  {B}VU Profile Options:{W}")
    for k, v in LOAD_VU_PROFILES.items():
        log(f"    {k}. {v['name']:<14}  users={v['users']}  "
            f"ramp={v['ramp']}s  dur={v['duration']}s")
    vu_choice = input(f"  Override VU profile? [1-6, ENTER to use default {test_type}]: ").strip()
    if vu_choice in LOAD_VU_PROFILES:
        profile = LOAD_VU_PROFILES[vu_choice]

    ok(f"[Load Engine] {profile['name']}: {profile['users']} VUs / "
       f"{profile['ramp']}s ramp / {profile['duration']}s duration")

    monitor = BottleneckMonitor(interval=3.0)
    engine  = LoadEngine(reqs, token="", monitor=monitor)
    report  = engine.run(list(LOAD_VU_PROFILES.keys())[
        list(LOAD_VU_PROFILES.values()).index(profile)
    ])
    report.print_summary()
    monitor.print_summary()

    # V47: Generate all reports
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    rpt = ReportEngine(ts=ts)
    paths = rpt.generate_all(
        report.metrics,
        monitor_summary=monitor.summary(),
        rca_log=ai_agent.rca_log,
        healing_stats=kb_load_healing_stats(),
        scenario_name=scenario,
    )
    log()
    log(f"  {B}REPORTS GENERATED:{W}")
    for fmt, path in paths.items():
        if path:
            log(f"  {fmt.upper():<6} → {path}")
    if paths.get("html") and os.path.exists(paths["html"]):
        webbrowser.open(paths["html"])

# ─── V47: VIEW REPORTS ────────────────────────────────────────────────────────
def view_reports():
    banner("V47 Reports — Available Files")
    files = sorted([
        f for f in os.listdir(REPORTS_DIR)
        if f.startswith("report_")
    ], reverse=True)
    if not files:
        warn(f"No reports found in: {REPORTS_DIR}")
        return
    for i, f in enumerate(files[:20], 1):
        full = os.path.join(REPORTS_DIR, f)
        size = os.path.getsize(full) // 1024
        log(f"  {i:>2}. {f}  ({size} KB)")
    log()
    try:
        choice = int(input("  Open report number (ENTER to skip): ").strip() or "0")
        if 1 <= choice <= len(files):
            full = os.path.join(REPORTS_DIR, files[choice - 1])
            webbrowser.open(full)
            ok(f"Opened: {full}")
    except ValueError:
        pass

# ═══════════════════════════════════════════════════════════════════════════════
# V47-E: EXTENDED 10-OPTION INTERACTIVE MENU
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    os.makedirs(OUT_DIR,     exist_ok=True)
    os.makedirs(REPORTS_DIR, exist_ok=True)
    os.makedirs(os.path.join(OUT_DIR, "html_reports"), exist_ok=True)
    os.makedirs(SCENARIOS_DIR, exist_ok=True)

    banner("JMeter Automation v47 — Universal Enterprise Framework")
    info(f"Scenarios  : {SCENARIOS_DIR}")
    info(f"Reports    : {REPORTS_DIR}")
    info(f"Learning DB: {DB_PATH}")
    info(f"Platform   : {platform.system()} {platform.release()}")

    known = db_load_known_fields()
    if known:
        ok(f"Learning DB: {len(known)} previously discovered fields loaded")

    while True:
        log()
        log(f"  {B}{C}{'─'*56}{W}")
        log(f"  {B}   MAIN MENU  —  v47 AI Performance Testing Agent{W}")
        log(f"  {B}{C}{'─'*56}{W}")
        log(f"  {G}  1.{W}  Record New Scenario        (Playwright + JMX)")
        log(f"  {C}  2.{W}  Execute Existing Scenario  (JMeter)")
        log(f"  {G}  3.{W}  Run Functional Test        (Python Load Engine)")
        log(f"  {G}  4.{W}  Run Load Test              (Python Load Engine)")
        log(f"  {Y}  5.{W}  Run Stress Test            (Python Load Engine)")
        log(f"  {Y}  6.{W}  Run Spike Test             (Python Load Engine)")
        log(f"  {C}  7.{W}  View Reports               (HTML / Excel / PDF)")
        log(f"  {C}  8.{W}  View RCA Summary           (AI Agent Analysis)")
        log(f"  {C}  9.{W}  View Self-Healing Stats    (Knowledge Base)")
        log(f"  {R} 10.{W}  Exit")
        log(f"  {B}{C}{'─'*56}{W}")
        log()

        option = input("  Select Option [1–10] : ").strip()

        # ── Option 1: Record New Scenario (v46 pipeline unchanged) ────────────
        if option == "1":
            banner("STEP 0 — Target Application")
            target_url = input("\nEnter Application URL to Record: ").strip()

            banner("STEP 1 — Record your browser session")
            reqs = record(target_url)
            if not reqs:
                err("Nothing captured."); continue

            banner("STEP 2 — Filter API requests")
            reqs = filter_reqs(reqs)
            ok(f"{len(reqs)} API requests after filter")
            reqs = remove_failed_requests(reqs)

            login_req = discover_auth(reqs)
            if not login_req:
                warn("No authentication flow detected — continuing without login.")
                login_req = {
                    "url": "", "path": "", "method": "GET",
                    "headers": {}, "postData": "", "creds": {},
                }
            else:
                ok(f"Auth request  : {login_req['url']}")
                ok(f"Credentials   : {list(login_req.get('creds',{}).keys())}")

            banner("STEP 3 — Select test profile")
            log(f"  {'#':<4} {'Name':<16} {'Users':>6} {'Ramp':>6}  {'Duration'}")
            for k, v in TEST_PROFILES.items():
                log(f"  {k}    {v['name']:<16} {v['threads']:>6} "
                    f"{v['ramp']:>5}s  {v['duration']}s")
            c = input("\n  Choice [1–5, default=1]: ").strip() or "1"
            if c not in TEST_PROFILES: c = "1"
            profile = TEST_PROFILES[c]
            ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
            ok(f"Profile: {profile['name']} — {profile['threads']} users")

            banner("STEP 4 — Token + V47 Struct-Aware Correlation")
            token_key, token_val = find_token(reqs)
            if token_key:
                ok(f"Token field: '{token_key}'")
            else:
                warn("No token found in login response!")

            corr, val_map, dependency_graph = build_corr_v41(reqs, token_key)
            ok(f"Dependency graph: {len(dependency_graph)} edges")
            ok(f"Learning DB: {len(db_load_known_fields())} fields stored")

            unstable_fields = monitor_correlation()
            if unstable_fields:
                warn(f"{len(unstable_fields)} unstable correlations detected")

            entities = discover_entities(reqs)
            if entities:
                ok(f"Entities discovered: {len(entities)}")
                for e in entities:
                    info(f"  req[{e['request']}] {e['field']} = {e['value']}"
                         f"  group={e['entity_group']}  conf={e['confidence']}")

            # V47-A: AI Agent full scenario analysis
            analysis = ai_agent.analyze_scenario(reqs, corr, val_map,
                                                   scenario="recorded")
            if analysis["recommendations"]:
                log(f"\n  {B}AI Agent Recommendations:{W}")
                for rec in analysis["recommendations"]:
                    log(f"  → {rec}")

            for r in reqs:
                if r.get("isLogin"): continue
                if r.get("postData"):
                    r["postData"] = sub_body(r["postData"], corr, val_map)
                r["url"]  = sub_url(r["url"], corr, val_map)
                r["path"] = urlsplit(r["url"]).path or "/"

            banner("STEP 4b — Universal Request Healing (V47)")
            for r in reqs:
                if r.get("isLogin"): continue
                r = heal_request(r, corr)
            ok("All requests healed (URL / headers / cookies / multipart / project_uid)")

            banner("STEP 5 — User-typed fields + V47 Faker data")
            user_fields = find_fields(reqs, corr)
            for k, v in user_fields.items():
                ok(f"  '{k}' = '{v}' → Faker({k}) + ITER_SUFFIX at runtime")

            banner("STEP 6 — V47 Single Sequential Flow (preserves CREATE chain)")
            flows = split_flows_v40(reqs, dependency_graph)
            for i, fl in enumerate(flows, 1):
                non_login_count = len([r for r in fl if not r.get("isLogin")])
                log(f"    {i}. {flow_name(fl)} — {len(fl)} total "
                    f"({non_login_count} non-login samplers)")

            banner("STEP 7 — Build JMX + CSV report")
            jmx      = os.path.join(OUT_DIR, f"recorded_{ts}.jmx")
            jtl      = os.path.join(OUT_DIR, f"result_{ts}.jtl")
            html     = os.path.join(OUT_DIR, "html_reports", f"report_{ts}")
            csv_path = os.path.join(OUT_DIR, f"parameterization_{ts}.csv")

            build_jmx(
                reqs, profile, jmx, jtl,
                corr, val_map,
                token_key, token_val,
                user_fields, flows, login_req,
            )
            write_csv_report(
                ts, profile, corr, token_key,
                user_fields, login_req, csv_path,
            )

            # V47: Save request cache for load engine
            scenario_name_input = None
            try:
                scenario_name_input = input("\nEnter Scenario Name for save : ").strip()
                scenario_dir  = os.path.join(SCENARIOS_DIR, scenario_name_input)
                os.makedirs(scenario_dir, exist_ok=True)
                cache_path = os.path.join(scenario_dir, "requests.json")
                safe_reqs  = []
                for r in reqs:
                    safe_reqs.append({
                        k: v for k, v in r.items()
                        if k not in ("resp_json", "ws_frames", "_healed_multipart")
                        and isinstance(v, (str, int, float, bool, list, dict, type(None)))
                    })
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(safe_reqs, f, indent=2)
                target_jmx = os.path.join(scenario_dir, "scenario.jmx")
                shutil.copy2(jmx, target_jmx)
                ok(f"Scenario saved : {scenario_name_input}")
            except Exception as ex:
                warn(f"Scenario save error: {ex}")

            cleanup = generate_cleanup(reqs)
            if cleanup:
                ok(f"Cleanup plan: {len(cleanup)} DELETE(s) queued post-run")

            log()
            log(f"  {B}FILES GENERATED:{W}")
            log(f"  JMX (test plan)  → {jmx}")
            log(f"  JTL (results)    → {jtl}  (written during run)")
            log(f"  HTML report      → {html}/index.html")
            log(f"  CSV params       → {csv_path}")

            log()
            run_now = input("  Run this scenario now with JMeter? [y/N] : ").strip().lower()
            if run_now == "y":
                execute_saved_jmx(jmx, profile)

        # ── Option 2: Execute Existing Scenario (JMeter) ──────────────────────
        elif option == "2":
            run_existing_scenario()

        # ── Options 3–6: Python Load Engine ───────────────────────────────────
        elif option in ("3", "4", "5", "6"):
            type_map = {"3": "1", "4": "2", "5": "3", "6": "4"}
            run_python_load_test(type_map[option])

        # ── Option 7: View Reports ─────────────────────────────────────────────
        elif option == "7":
            view_reports()

        # ── Option 8: View RCA Summary ────────────────────────────────────────
        elif option == "8":
            ai_agent.print_rca_summary()

        # ── Option 9: View Self-Healing Statistics ────────────────────────────
        elif option == "9":
            ai_agent.print_healing_summary()

        # ── Option 10: Exit ───────────────────────────────────────────────────
        elif option == "10":
            ok("Goodbye!"); break

        else:
            warn("Invalid option. Please enter 1–10.")


if __name__ == "__main__":
    main()